import json
import math
import numpy as np
import BasicStats
import SingleMap
import Operators
from openpyxl import load_workbook


#File Path Location of Excel Sheet
Excel = '/home/gabe/Documents/MatchStats/main/Stats.xlsx'
#Load the workbook into read cells
workbook = load_workbook(filename=Excel)

#Header of 'Operator Stats' page in Excel
OperatorHeader = ['Username', 'Most Played Atk Op', '# of Rounds', 'Most Played Def Op', '# of Rounds','Highest Rating Atk Op', 'Rating', 'Highest Rating Def Op', 'Rating','Mute Rating', 'Mute Kills', 'Mute Deaths', 'Mute EntryKills', 'Mute EntryDeaths', 'Mute KOST', 'Mute HS', 'Mute MKills', 'Mute Trades', 'Mute Clutch', 'Mute Plants', 'Mute Rounds', 'Smoke Rating', 'Smoke Kills', 'Smoke Deaths', 'Smoke EntryKills', 'Smoke EntryDeaths', 'Smoke KOST', 'Smoke HS', 'Smoke MKills', 'Smoke Trades', 'Smoke Clutch', 'Smoke Plants', 'Smoke Rounds', 'Castle Rating', 'Castle Kills', 'Castle Deaths', 'Castle EntryKills', 'Castle EntryDeaths', 'Castle KOST', 'Castle HS', 'Castle MKills', 'Castle Trades', 'Castle Clutch', 'Castle Plants', 'Castle Rounds', 'Pulse Rating', 'Pulse Kills', 'Pulse Deaths', 'Pulse EntryKills', 'Pulse EntryDeaths', 'Pulse KOST', 'Pulse HS', 'Pulse MKills', 'Pulse Trades', 'Pulse Clutch', 'Pulse Plants', 'Pulse Rounds', 'Doc Rating', 'Doc Kills', 'Doc Deaths', 'Doc EntryKills', 'Doc EntryDeaths', 'Doc KOST', 'Doc HS', 'Doc MKills', 'Doc Trades', 'Doc Clutch', 'Doc Plants', 'Doc Rounds', 'Rook Rating', 'Rook Kills', 'Rook Deaths', 'Rook EntryKills', 'Rook EntryDeaths', 'Rook KOST', 'Rook HS', 'Rook MKills', 'Rook Trades', 'Rook Clutch', 'Rook Plants', 'Rook Rounds', 'Jager Rating', 'Jager Kills', 'Jager Deaths', 'Jager EntryKills', 'Jager EntryDeaths', 'Jager KOST', 'Jager HS', 'Jager MKills', 'Jager Trades', 'Jager Clutch', 'Jager Plants', 'Jager Rounds', 'Bandit Rating', 'Bandit Kills', 'Bandit Deaths', 'Bandit EntryKills', 'Bandit EntryDeaths', 'Bandit KOST', 'Bandit HS', 'Bandit MKills', 'Bandit Trades', 'Bandit Clutch', 'Bandit Plants', 'Bandit Rounds', 'Tachanka Rating', 'Tachanka Kills', 'Tachanka Deaths', 'Tachanka EntryKills', 'Tachanka EntryDeaths', 'Tachanka KOST', 'Tachanka HS', 'Tachanka MKills', 'Tachanka Trades', 'Tachanka Clutch', 'Tachanka Plants', 'Tachanka Rounds', 'Kapkan Rating', 'Kapkan Kills', 'Kapkan Deaths', 'Kapkan EntryKills', 'Kapkan EntryDeaths', 'Kapkan KOST', 'Kapkan HS', 'Kapkan MKills', 'Kapkan Trades', 'Kapkan Clutch', 'Kapkan Plants', 'Kapkan Rounds', 'Frost Rating', 'Frost Kills', 'Frost Deaths', 'Frost EntryKills', 'Frost EntryDeaths', 'Frost KOST', 'Frost HS', 'Frost MKills', 'Frost Trades', 'Frost Clutch', 'Frost Plants', 'Frost Rounds', 'Valkyrie Rating', 'Valkyrie Kills', 'Valkyrie Deaths', 'Valkyrie EntryKills', 'Valkyrie EntryDeaths', 'Valkyrie KOST', 'Valkyrie HS', 'Valkyrie MKills', 'Valkyrie Trades', 'Valkyrie Clutch', 'Valkyrie Plants', 'Valkyrie Rounds', 'Caveira Rating', 'Caveira Kills', 'Caveira Deaths', 'Caveira EntryKills', 'Caveira EntryDeaths', 'Caveira KOST', 'Caveira HS', 'Caveira MKills', 'Caveira Trades', 'Caveira Clutch', 'Caveira Plants', 'Caveira Rounds', 'Echo Rating', 'Echo Kills', 'Echo Deaths', 'Echo EntryKills', 'Echo EntryDeaths', 'Echo KOST', 'Echo HS', 'Echo MKills', 'Echo Trades', 'Echo Clutch', 'Echo Plants', 'Echo Rounds', 'Mira Rating', 'Mira Kills', 'Mira Deaths', 'Mira EntryKills', 'Mira EntryDeaths', 'Mira KOST', 'Mira HS', 'Mira MKills', 'Mira Trades', 'Mira Clutch', 'Mira Plants', 'Mira Rounds', 'Lesion Rating', 'Lesion Kills', 'Lesion Deaths', 'Lesion EntryKills', 'Lesion EntryDeaths', 'Lesion KOST', 'Lesion HS', 'Lesion MKills', 'Lesion Trades', 'Lesion Clutch', 'Lesion Plants', 'Lesion Rounds', 'Ela Rating', 'Ela Kills', 'Ela Deaths', 'Ela EntryKills', 'Ela EntryDeaths', 'Ela KOST', 'Ela HS', 'Ela MKills', 'Ela Trades', 'Ela Clutch', 'Ela Plants', 'Ela Rounds', 'Vigil Rating', 'Vigil Kills', 'Vigil Deaths', 'Vigil EntryKills', 'Vigil EntryDeaths', 'Vigil KOST', 'Vigil HS', 'Vigil MKills', 'Vigil Trades', 'Vigil Clutch', 'Vigil Plants', 'Vigil Rounds', 'Alibi Rating', 'Alibi Kills', 'Alibi Deaths', 'Alibi EntryKills', 'Alibi EntryDeaths', 'Alibi KOST', 'Alibi HS', 'Alibi MKills', 'Alibi Trades', 'Alibi Clutch', 'Alibi Plants', 'Alibi Rounds', 'Maestro Rating', 'Maestro Kills', 'Maestro Deaths', 'Maestro EntryKills', 'Maestro EntryDeaths', 'Maestro KOST', 'Maestro HS', 'Maestro MKills', 'Maestro Trades', 'Maestro Clutch', 'Maestro Plants', 'Maestro Rounds', 'Clash Rating', 'Clash Kills', 'Clash Deaths', 'Clash EntryKills', 'Clash EntryDeaths', 'Clash KOST', 'Clash HS', 'Clash MKills', 'Clash Trades', 'Clash Clutch', 'Clash Plants', 'Clash Rounds', 'Kaid Rating', 'Kaid Kills', 'Kaid Deaths', 'Kaid EntryKills', 'Kaid EntryDeaths', 'Kaid KOST', 'Kaid HS', 'Kaid MKills', 'Kaid Trades', 'Kaid Clutch', 'Kaid Plants', 'Kaid Rounds', 'Mozzie Rating', 'Mozzie Kills', 'Mozzie Deaths', 'Mozzie EntryKills', 'Mozzie EntryDeaths', 'Mozzie KOST', 'Mozzie HS', 'Mozzie MKills', 'Mozzie Trades', 'Mozzie Clutch', 'Mozzie Plants', 'Mozzie Rounds', 'Warden Rating', 'Warden Kills', 'Warden Deaths', 'Warden EntryKills', 'Warden EntryDeaths', 'Warden KOST', 'Warden HS', 'Warden MKills', 'Warden Trades', 'Warden Clutch', 'Warden Plants', 'Warden Rounds', 'Goyo Rating', 'Goyo Kills', 'Goyo Deaths', 'Goyo EntryKills', 'Goyo EntryDeaths', 'Goyo KOST', 'Goyo HS', 'Goyo MKills', 'Goyo Trades', 'Goyo Clutch', 'Goyo Plants', 'Goyo Rounds', 'Wamai Rating', 'Wamai Kills', 'Wamai Deaths', 'Wamai EntryKills', 'Wamai EntryDeaths', 'Wamai KOST', 'Wamai HS', 'Wamai MKills', 'Wamai Trades', 'Wamai Clutch', 'Wamai Plants', 'Wamai Rounds', 'Oryx Rating', 'Oryx Kills', 'Oryx Deaths', 'Oryx EntryKills', 'Oryx EntryDeaths', 'Oryx KOST', 'Oryx HS', 'Oryx MKills', 'Oryx Trades', 'Oryx Clutch', 'Oryx Plants', 'Oryx Rounds', 'Melusi Rating', 'Melusi Kills', 'Melusi Deaths', 'Melusi EntryKills', 'Melusi EntryDeaths', 'Melusi KOST', 'Melusi HS', 'Melusi MKills', 'Melusi Trades', 'Melusi Clutch', 'Melusi Plants', 'Melusi Rounds', 'Aruni Rating', 'Aruni Kills', 'Aruni Deaths', 'Aruni EntryKills', 'Aruni EntryDeaths', 'Aruni KOST', 'Aruni HS', 'Aruni MKills', 'Aruni Trades', 'Aruni Clutch', 'Aruni Plants', 'Aruni Rounds', 'Thunderbird Rating', 'Thunderbird Kills', 'Thunderbird Deaths', 'Thunderbird EntryKills', 'Thunderbird EntryDeaths', 'Thunderbird KOST', 'Thunderbird HS', 'Thunderbird MKills', 'Thunderbird Trades', 'Thunderbird Clutch', 'Thunderbird Plants', 'Thunderbird Rounds', 'Thorn Rating', 'Thorn Kills', 'Thorn Deaths', 'Thorn EntryKills', 'Thorn EntryDeaths', 'Thorn KOST', 'Thorn HS', 'Thorn MKills', 'Thorn Trades', 'Thorn Clutch', 'Thorn Plants', 'Thorn Rounds', 'Azami Rating', 'Azami Kills', 'Azami Deaths', 'Azami EntryKills', 'Azami EntryDeaths', 'Azami KOST', 'Azami HS', 'Azami MKills', 'Azami Trades', 'Azami Clutch', 'Azami Plants', 'Azami Rounds', 'Solis Rating', 'Solis Kills', 'Solis Deaths', 'Solis EntryKills', 'Solis EntryDeaths', 'Solis KOST', 'Solis HS', 'Solis MKills', 'Solis Trades', 'Solis Clutch', 'Solis Plants', 'Solis Rounds', 'Fenrir Rating', 'Fenrir Kills', 'Fenrir Deaths', 'Fenrir EntryKills', 'Fenrir EntryDeaths', 'Fenrir KOST', 'Fenrir HS', 'Fenrir MKills', 'Fenrir Trades', 'Fenrir Clutch', 'Fenrir Plants', 'Fenrir Rounds', 'Sledge Rating', 'Sledge Kills', 'Sledge Deaths', 'Sledge EntryKills', 'Sledge EntryDeaths', 'Sledge KOST', 'Sledge HS', 'Sledge MKills', 'Sledge Trades', 'Sledge Clutch', 'Sledge Plants', 'Sledge Rounds', 'Thatcher Rating', 'Thatcher Kills', 'Thatcher Deaths', 'Thatcher EntryKills', 'Thatcher EntryDeaths', 'Thatcher KOST', 'Thatcher HS', 'Thatcher MKills', 'Thatcher Trades', 'Thatcher Clutch', 'Thatcher Plants', 'Thatcher Rounds', 'Ash Rating', 'Ash Kills', 'Ash Deaths', 'Ash EntryKills', 'Ash EntryDeaths', 'Ash KOST', 'Ash HS', 'Ash MKills', 'Ash Trades', 'Ash Clutch', 'Ash Plants', 'Ash Rounds', 'Thermite Rating', 'Thermite Kills', 'Thermite Deaths', 'Thermite EntryKills', 'Thermite EntryDeaths', 'Thermite KOST', 'Thermite HS', 'Thermite MKills', 'Thermite Trades', 'Thermite Clutch', 'Thermite Plants', 'Thermite Rounds', 'Montagne Rating', 'Montagne Kills', 'Montagne Deaths', 'Montagne EntryKills', 'Montagne EntryDeaths', 'Montagne KOST', 'Montagne HS', 'Montagne MKills', 'Montagne Trades', 'Montagne Clutch', 'Montagne Plants', 'Montagne Rounds', 'Twitch Rating', 'Twitch Kills', 'Twitch Deaths', 'Twitch EntryKills', 'Twitch EntryDeaths', 'Twitch KOST', 'Twitch HS', 'Twitch MKills', 'Twitch Trades', 'Twitch Clutch', 'Twitch Plants', 'Twitch Rounds', 'Blitz Rating', 'Blitz Kills', 'Blitz Deaths', 'Blitz EntryKills', 'Blitz EntryDeaths', 'Blitz KOST', 'Blitz HS', 'Blitz MKills', 'Blitz Trades', 'Blitz Clutch', 'Blitz Plants', 'Blitz Rounds', 'IQ Rating', 'IQ Kills', 'IQ Deaths', 'IQ EntryKills', 'IQ EntryDeaths', 'IQ KOST', 'IQ HS', 'IQ MKills', 'IQ Trades', 'IQ Clutch', 'IQ Plants', 'IQ Rounds', 'Fuze Rating', 'Fuze Kills', 'Fuze Deaths', 'Fuze EntryKills', 'Fuze EntryDeaths', 'Fuze KOST', 'Fuze HS', 'Fuze MKills', 'Fuze Trades', 'Fuze Clutch', 'Fuze Plants', 'Fuze Rounds', 'Glaz Rating', 'Glaz Kills', 'Glaz Deaths', 'Glaz EntryKills', 'Glaz EntryDeaths', 'Glaz KOST', 'Glaz HS', 'Glaz MKills', 'Glaz Trades', 'Glaz Clutch', 'Glaz Plants', 'Glaz Rounds', 'Buck Rating', 'Buck Kills', 'Buck Deaths', 'Buck EntryKills', 'Buck EntryDeaths', 'Buck KOST', 'Buck HS', 'Buck MKills', 'Buck Trades', 'Buck Clutch', 'Buck Plants', 'Buck Rounds', 'Blackbeard Rating', 'Blackbeard Kills', 'Blackbeard Deaths', 'Blackbeard EntryKills', 'Blackbeard EntryDeaths', 'Blackbeard KOST', 'Blackbeard HS', 'Blackbeard MKills', 'Blackbeard Trades', 'Blackbeard Clutch', 'Blackbeard Plants', 'Blackbeard Rounds', 'Capitao Rating', 'Capitao Kills', 'Capitao Deaths', 'Capitao EntryKills', 'Capitao EntryDeaths', 'Capitao KOST', 'Capitao HS', 'Capitao MKills', 'Capitao Trades', 'Capitao Clutch', 'Capitao Plants', 'Capitao Rounds', 'Hibana Rating', 'Hibana Kills', 'Hibana Deaths', 'Hibana EntryKills', 'Hibana EntryDeaths', 'Hibana KOST', 'Hibana HS', 'Hibana MKills', 'Hibana Trades', 'Hibana Clutch', 'Hibana Plants', 'Hibana Rounds', 'Jackal Rating', 'Jackal Kills', 'Jackal Deaths', 'Jackal EntryKills', 'Jackal EntryDeaths', 'Jackal KOST', 'Jackal HS', 'Jackal MKills', 'Jackal Trades', 'Jackal Clutch', 'Jackal Plants', 'Jackal Rounds', 'Ying Rating', 'Ying Kills', 'Ying Deaths', 'Ying EntryKills', 'Ying EntryDeaths', 'Ying KOST', 'Ying HS', 'Ying MKills', 'Ying Trades', 'Ying Clutch', 'Ying Plants', 'Ying Rounds', 'Zofia Rating', 'Zofia Kills', 'Zofia Deaths', 'Zofia EntryKills', 'Zofia EntryDeaths', 'Zofia KOST', 'Zofia HS', 'Zofia MKills', 'Zofia Trades', 'Zofia Clutch', 'Zofia Plants', 'Zofia Rounds', 'Dokkaebi Rating', 'Dokkaebi Kills', 'Dokkaebi Deaths', 'Dokkaebi EntryKills', 'Dokkaebi EntryDeaths', 'Dokkaebi KOST', 'Dokkaebi HS', 'Dokkaebi MKills', 'Dokkaebi Trades', 'Dokkaebi Clutch', 'Dokkaebi Plants', 'Dokkaebi Rounds', 'Finka Rating', 'Finka Kills', 'Finka Deaths', 'Finka EntryKills', 'Finka EntryDeaths', 'Finka KOST', 'Finka HS', 'Finka MKills', 'Finka Trades', 'Finka Clutch', 'Finka Plants', 'Finka Rounds', 'Lion Rating', 'Lion Kills', 'Lion Deaths', 'Lion EntryKills', 'Lion EntryDeaths', 'Lion KOST', 'Lion HS', 'Lion MKills', 'Lion Trades', 'Lion Clutch', 'Lion Plants', 'Lion Rounds', 'Maverick Rating', 'Maverick Kills', 'Maverick Deaths', 'Maverick EntryKills', 'Maverick EntryDeaths', 'Maverick KOST', 'Maverick HS', 'Maverick MKills', 'Maverick Trades', 'Maverick Clutch', 'Maverick Plants', 'Maverick Rounds', 'Nomad Rating', 'Nomad Kills', 'Nomad Deaths', 'Nomad EntryKills', 'Nomad EntryDeaths', 'Nomad KOST', 'Nomad HS', 'Nomad MKills', 'Nomad Trades', 'Nomad Clutch', 'Nomad Plants', 'Nomad Rounds', 'Gridlock Rating', 'Gridlock Kills', 'Gridlock Deaths', 'Gridlock EntryKills', 'Gridlock EntryDeaths', 'Gridlock KOST', 'Gridlock HS', 'Gridlock MKills', 'Gridlock Trades', 'Gridlock Clutch', 'Gridlock Plants', 'Gridlock Rounds', 'Nokk Rating', 'Nokk Kills', 'Nokk Deaths', 'Nokk EntryKills', 'Nokk EntryDeaths', 'Nokk KOST', 'Nokk HS', 'Nokk MKills', 'Nokk Trades', 'Nokk Clutch', 'Nokk Plants', 'Nokk Rounds', 'Amaru Rating', 'Amaru Kills', 'Amaru Deaths', 'Amaru EntryKills', 'Amaru EntryDeaths', 'Amaru KOST', 'Amaru HS', 'Amaru MKills', 'Amaru Trades', 'Amaru Clutch', 'Amaru Plants', 'Amaru Rounds', 'Kali Rating', 'Kali Kills', 'Kali Deaths', 'Kali EntryKills', 'Kali EntryDeaths', 'Kali KOST', 'Kali HS', 'Kali MKills', 'Kali Trades', 'Kali Clutch', 'Kali Plants', 'Kali Rounds', 'Iana Rating', 'Iana Kills', 'Iana Deaths', 'Iana EntryKills', 'Iana EntryDeaths', 'Iana KOST', 'Iana HS', 'Iana MKills', 'Iana Trades', 'Iana Clutch', 'Iana Plants', 'Iana Rounds', 'Ace Rating', 'Ace Kills', 'Ace Deaths', 'Ace EntryKills', 'Ace EntryDeaths', 'Ace KOST', 'Ace HS', 'Ace MKills', 'Ace Trades', 'Ace Clutch', 'Ace Plants', 'Ace Rounds', 'Zero Rating', 'Zero Kills', 'Zero Deaths', 'Zero EntryKills', 'Zero EntryDeaths', 'Zero KOST', 'Zero HS', 'Zero MKills', 'Zero Trades', 'Zero Clutch', 'Zero Plants', 'Zero Rounds', 'Flores Rating', 'Flores Kills', 'Flores Deaths', 'Flores EntryKills', 'Flores EntryDeaths', 'Flores KOST', 'Flores HS', 'Flores MKills', 'Flores Trades', 'Flores Clutch', 'Flores Plants', 'Flores Rounds', 'Osa Rating', 'Osa Kills', 'Osa Deaths', 'Osa EntryKills', 'Osa EntryDeaths', 'Osa KOST', 'Osa HS', 'Osa MKills', 'Osa Trades', 'Osa Clutch', 'Osa Plants', 'Osa Rounds', 'Sens Rating', 'Sens Kills', 'Sens Deaths', 'Sens EntryKills', 'Sens EntryDeaths', 'Sens KOST', 'Sens HS', 'Sens MKills', 'Sens Trades', 'Sens Clutch', 'Sens Plants', 'Sens Rounds', 'Grim Rating', 'Grim Kills', 'Grim Deaths', 'Grim EntryKills', 'Grim EntryDeaths', 'Grim KOST', 'Grim HS', 'Grim MKills', 'Grim Trades', 'Grim Clutch', 'Grim Plants', 'Grim Rounds', 'Brava Rating', 'Brava Kills', 'Brava Deaths', 'Brava EntryKills', 'Brava EntryDeaths', 'Brava KOST', 'Brava HS', 'Brava MKills', 'Brava Trades', 'Brava Clutch', 'Brava Plants', 'Brava Rounds']
# Select the desired sheet
sheet_name = 'Stats'
Op_sheet_name = 'Operator Stats'
sheet = workbook[sheet_name]
sheet1 = workbook[Op_sheet_name]
#Define variables
ExcelUsername = []
ExcelRating = []
ExcelKills = []
ExcelDeaths = []
ExcelKD = []
ExcelEK = []
ExcelED = []
ExcelEntry = []
ExcelKOST = []
ExcelKPR = []
ExcelSRV = []
ExcelMKills = []
ExcelTrade = []
ExcelClutch = []
ExcelPlants = []
ExcelHS = []
ExcelAtk = []
ExcelDef = []
ExcelRound = []
ExcelColUsername = 'A'
ExcelColRating = 'B'
ExcelColKills = 'C' 
ExcelColDeaths = 'D'
ExcelColKD = 'E'
ExcelColEK = 'F'
ExcelColED = 'G'
ExcelColEntry = 'H'
ExcelColKOST = 'I'
ExcelColKPR = 'J'
ExcelColSRV = 'K'
ExcelColMKills = 'L'
ExcelColTrade = 'M'
ExcelColClutch = 'N'
ExcelColPlants = 'O'
ExcelColHS = 'P'
ExcelColRound = 'Q'

#Array of the column names to make it easier to access individual columns below
ExcelCols = [ExcelColUsername,ExcelColRating, ExcelColKills,ExcelColDeaths,ExcelColKD,ExcelColEK,ExcelColED,ExcelColEntry,ExcelColKOST,ExcelColKPR,ExcelColSRV,ExcelColMKills,ExcelColTrade,ExcelColClutch,ExcelColPlants,ExcelColHS,ExcelColRound]

#For each column, save all of data to an array
#Starts at the second cell because the Header is written in the Excel File
for cell in sheet[ExcelCols[0]]:
    ExcelUsername.append(cell.value)
ExcelUsername = ExcelUsername[1:]
for cell in sheet[ExcelCols[1]]:
    ExcelRating.append(cell.value)
ExcelRating = ExcelRating[1:]
for cell in sheet[ExcelCols[2]]:
    ExcelKills.append(cell.value)
ExcelKills = ExcelKills[1:]
for cell in sheet[ExcelCols[3]]:
    ExcelDeaths.append(cell.value)
ExcelDeaths = ExcelDeaths[1:]
for cell in sheet[ExcelCols[4]]:
    ExcelKD.append(cell.value)
ExcelKD = ExcelKD[1:]
for cell in sheet[ExcelCols[5]]:
    ExcelEK.append(cell.value)
ExcelEK = ExcelEK[1:]
for cell in sheet[ExcelCols[6]]:
    ExcelED.append(cell.value)
ExcelED = ExcelED[1:]
for cell in sheet[ExcelCols[7]]:
    ExcelEntry.append(cell.value)
ExcelEntry = ExcelEntry[1:]
for cell in sheet[ExcelCols[8]]:
    ExcelKOST.append(cell.value)
ExcelKOST = ExcelKOST[1:]
for cell in sheet[ExcelCols[9]]:
    ExcelKPR.append(cell.value)
ExcelKPR = ExcelKPR[1:]
for cell in sheet[ExcelCols[10]]:
    ExcelSRV.append(cell.value)
ExcelSRV = ExcelSRV[1:]
for cell in sheet[ExcelCols[11]]:
    ExcelMKills.append(cell.value)
ExcelMKills = ExcelMKills[1:]
for cell in sheet[ExcelCols[12]]:
    ExcelTrade.append(cell.value)
ExcelTrade = ExcelTrade[1:]
for cell in sheet[ExcelCols[13]]:
    ExcelClutch.append(cell.value)
ExcelClutch = ExcelClutch[1:]
for cell in sheet[ExcelCols[14]]:
    ExcelPlants.append(cell.value)
ExcelPlants = ExcelPlants[1:]
for cell in sheet[ExcelCols[15]]:
    ExcelHS.append(cell.value)
ExcelHS = ExcelHS[1:]
for cell in sheet[ExcelCols[16]]:
    ExcelRound.append(cell.value)
ExcelRound = ExcelRound[1:]


#Takes in the length of the ExcelUsername for 'Operator Stats' to read that many lines from the Operator Stats and read down the row
#Creates a 2d array of width of array of ExcelUsername and Length of # of players
Op_array = []

if len(ExcelUsername) != 0:
    for i in range(len(ExcelUsername)):
        temparr = []
        for cell in sheet1[2+i]:
            temparr.append(cell.value)
        Op_array.append(temparr)










timeToTrade = 10 #seconds
OpNumbers = 68 #number of operators
# <= 33 is defender, > 33 is attacker

#Definiton of rating system, same as used in basic stats, !!SHOULD BECOME UNIFORM!!
def ratingSys(rKills, rKD, rMK, rEntry, rPlants, rClutch, rKOST, rSRV, rRounds):
    rating = (rKD**2 + 0.4*(rKills) + 0.15*rMK)/rRounds + 0.75*(rEntry)/rRounds + (rPlants + rClutch)/rRounds + rKOST + rSRV/3
    return rating

        
#open the json file i am parsing
with open("/home/gabe/Documents/MatchStats/json_test/scrim6.json", 'r') as f:
    my_dict = json.load(f)

#function to parse data
[Users, Kills, Deaths, EKills, EDeaths, KOST, HS, MK, Trade, Clutch, Plant, Defuse, AttackerMain, DefenderMain, Rounds, OperatorKills, OperatorDeaths, OperatorEntryKills, OperatorEntryDeaths, OperatorKOST, OperatorHS, OperatorMKills, OperatorTrades, OperatorClutch, OperatorPlants, OperatorDefusal, OperatorRounds, MapPlayed, Team1Score, Team2Score, Team1Name, Team2Name] = SingleMap.singleMap(my_dict)

#use for loop as basic tool to print all player data, similar to siege GG
players = []

#Format output of function to what is necessary to the Basic Stats class
for i in range(len(Users)):
    #Intermediate values to make a potentially undefined value, or additional info necessary, like entry being subtracted
    #IE, if a person played 1 round, got 2 kills and didnt die, KD would be infinite if not reset
    OperatorEntry = np.zeros(OpNumbers)
    OperatorKD = np.zeros(OpNumbers)
    OperatorKOSTAmount = np.zeros(OpNumbers)
    OperatorKPR = np.zeros(OpNumbers)
    OperatorSRV = np.zeros(OpNumbers)
    KD = Kills[i]/Deaths[i]
    Entry = EKills[i]-EDeaths[i]
    KOSTAmount = KOST[i]/Rounds[i]
    KPR = Kills[i]/Rounds[i]
    SRV = (Rounds[i]-Deaths[i])/Rounds[i]
    #Repeats for number of operators as similar stats
    for j in range(OpNumbers):
        OperatorEntry[j] = OperatorEntryKills[i][j] - OperatorEntryDeaths[i][j]
        if OperatorDeaths[i][j] == 0:
            OperatorKD[j] = OperatorKills[i][j]
        else:
            OperatorKills[i][j]/OperatorDeaths[i][j]
        if OperatorRounds[i][j] == 0:
            OperatorKOSTAmount[j] = 0
            OperatorKPR[j] = 0
            OperatorSRV[j] = 0
        else:
            OperatorKOSTAmount[j] = OperatorKOST[i][j]/OperatorRounds[i][j]
            OperatorKPR[j] = OperatorKills[i][j]/OperatorRounds[i][j]
            OperatorSRV[j] = (OperatorRounds[i][j] - OperatorDeaths[i][j])/OperatorRounds[i][j]

    player = BasicStats.BasicStats(Users[i],Kills[i],Deaths[i],KD,EKills[i],EDeaths[i],Entry,KOSTAmount,KPR,SRV,MK[i],Trade[i],Clutch[i],Plant[i],Defuse[i],HS[i],AttackerMain[i], DefenderMain[i], Rounds[i],OperatorKills[i], OperatorDeaths[i],OperatorKD, OperatorEntryKills[i], OperatorEntryDeaths[i], OperatorEntry, OperatorKOSTAmount, OperatorKPR, OperatorSRV, OperatorMKills[i], OperatorTrades[i], OperatorClutch[i], OperatorPlants[i], OperatorDefusal[i], OperatorHS[i], OperatorRounds[i], MapPlayed, Team1Score, Team2Score, Team1Name, Team2Name)
    players.append(player)
    #Prints to terminal
    if(i==0):
        start = 1
    else:
        start = 0
    players[i].printIndivStat(start)
#Can be an outdated value
ExcelUserLists = len(ExcelUsername)
#For the number of users in the current match
for i in range(len(Users)):
    matching = 0
    matchingValue = 0
    #Checks these names against the list in Excel
    for j in range(len(ExcelUsername)):
        #If it matches, can confirm that they have played prior
        if Users[i] == Op_array[j][0]:
            matchingValue = j
            matching = matching + 1
    #If doesnt match, add the player to the list
    if matching == 0:
        array = []
        OperatorRating = []
        singleUser = []
        tempKD = 0
        #Same stopping of infinite values and replacing or undefined values
        for b in range(len(Operators.Operators)):
            if OperatorKills[i][b] == 0:
                tempHS = 0
            else:
                tempHS = OperatorHS[i][b]/OperatorKills[i][b]
            if OperatorDeaths[i][b] == 0:
                tempKD = OperatorKills[i][b]
            else:
                tempKD = OperatorKills[i][b]/OperatorDeaths[i][b]
            if OperatorRounds[i][b] == 0:
                tempKOST = 0
                tempSRV = 0
                tempRating = 0
            else:
                tempKOST = OperatorKOST[i][b]/OperatorRounds[i][b]
                tempSRV = (OperatorRounds[i][b] - OperatorDeaths[i][b])/OperatorRounds[i][b]
                tempRating = ratingSys(OperatorKills[i][b],tempKD, OperatorMKills[i][b], OperatorEntryKills[i][b] - OperatorEntryDeaths[i][b], OperatorPlants[i][b], OperatorClutch[i][b], tempKOST, tempSRV, OperatorRounds[i][b] )

            #Array to add to excel when all compiled
            array.append(round((tempRating),2))
            OperatorRating.append(round((tempRating),2))
            array.append(OperatorKills[i][b])
            array.append(OperatorDeaths[i][b])
            array.append(OperatorEntryKills[i][b])
            array.append(OperatorEntryDeaths[i][b])
            array.append(round(tempKOST,2))
            array.append(round(tempHS,2))
            array.append(OperatorMKills[i][b])
            array.append(OperatorTrades[i][b])
            array.append(OperatorClutch[i][b])
            array.append(OperatorPlants[i][b])
            array.append(OperatorRounds[i][b])
        singleUser.append(Users[i])
        #Within the Operators, looking at rounds for the operator plays and split between attackers and defenders
        OperatorDefRounds = OperatorRounds[i][0:33]
        OperatorDefRating = OperatorRating[0:33]
        OperatorAtkRounds = OperatorRounds[i][34:]
        OperatorAtkRating = OperatorRating[34:]
        
        #Find the max of the Rounds and ratings, these arrays are of slightly different types, but calculate the same thing
        maxDefRounds = np.argmax(OperatorDefRounds)
        maxAtkRounds = np.argmax(OperatorAtkRounds)
        maxDefRating = max(OperatorDefRating)
        maxAtkRating = max(OperatorAtkRating)
        
        #Append these values to the user array for excel
        singleUser.append(Operators.OperatorsValues[34+maxAtkRounds])
        singleUser.append(Operators.OperatorRounds[i][34+maxAtkRounds])
        singleUser.append(Operators.OperatorsValues[maxDefRounds])
        singleUser.append(Operators.OperatorRounds[i][maxDefRounds])
        
        singleUser.append(Operators.OperatorsValues[34+OperatorAtkRating.index(maxAtkRating)])
        singleUser.append(OperatorRating[34+OperatorAtkRating.index(maxAtkRating)])
        singleUser.append(Operators.OperatorsValues[OperatorDefRating.index(maxDefRating)])
        singleUser.append(OperatorRating[OperatorDefRating.index(maxDefRating)])

        #Adds the operator values to the single User arrays
        for f in range(len(array)):
            singleUser.append(array[f])
        Op_array.append(singleUser)
    #This is the values when a player already exists in excel
    else:
        tempAtkRating = []
        tempDefRating = []
        tempAtkRounds = []
        tempDefRounds = []
        #Add values the operator array that will be put into the excel
        #Updates the operator values that already exist in the values
        #Offsets based on excel to access the correct data
        for d in range(len(SingleMap.Operators.Operators)):
            #This is checking headshots and headshot value
            if Op_array[matchingValue][10+12*d] == 0:
                Op_array[matchingValue][15+12*d] = 0
            else:
                Op_array[matchingValue][15+12*d] = round((Op_array[matchingValue][15+12*d]*Op_array[matchingValue][10+12*d] + OperatorHS[i][d])/(Op_array[matchingValue][10+12*d] + OperatorKills[i][d]),2)
            #Updates all values that are empirical
            Op_array[matchingValue][10+12*d] = Op_array[matchingValue][10+12*d] + OperatorKills[i][d]
            Op_array[matchingValue][11+12*d] = Op_array[matchingValue][11+12*d] + OperatorDeaths[i][d]
            Op_array[matchingValue][12+12*d] = Op_array[matchingValue][12+12*d] + OperatorEntryKills[i][d]
            Op_array[matchingValue][13+12*d] = Op_array[matchingValue][13+12*d] + OperatorEntryDeaths[i][d]
            Op_array[matchingValue][16+12*d] = Op_array[matchingValue][16+12*d] + OperatorMKills[i][d]
            Op_array[matchingValue][17+12*d] = Op_array[matchingValue][17+12*d] + OperatorTrades[i][d]
            Op_array[matchingValue][18+12*d] = Op_array[matchingValue][18+12*d] + OperatorClutch[i][d]
            Op_array[matchingValue][19+12*d] = Op_array[matchingValue][19+12*d] + OperatorPlants[i][d]
            Op_array[matchingValue][20+12*d] = Op_array[matchingValue][20+12*d] + OperatorRounds[i][d]
            #Calculates Operator KD
            if Op_array[matchingValue][11+12*d] == 0:
                tempKD = Op_array[matchingValue][10+12*d]
            else:
                tempKD = Op_array[matchingValue][10+12*d]/Op_array[matchingValue][11+12*d]
            #Calculates survival
            if Op_array[matchingValue][20+12*d] == 0:
                Op_array[matchingValue][14+12*d] = 0
                tempSRV = 0
                tempRating = 0
            else:
                Op_array[matchingValue][14+12*d] = round((Op_array[matchingValue][14+12*d]*Op_array[matchingValue][20+12*d] + OperatorKOST[i][d])/(Op_array[matchingValue][20+12*d] + OperatorRounds[i][d]),2)
                tempSRV = (Op_array[matchingValue][20+12*d] - Op_array[matchingValue][11+12*d])/Op_array[matchingValue][20+12*d]
                tempRating = round(ratingSys(Op_array[matchingValue][10+12*d], tempKD, Op_array[matchingValue][16+12*d], Op_array[matchingValue][12+12*d] - Op_array[matchingValue][13+12*d], Op_array[matchingValue][19+12*d], Op_array[matchingValue][18+12*d], Op_array[matchingValue][14+12*d], tempSRV, Op_array[matchingValue][20+12*d]),2)
            Op_array[matchingValue][9+12*d] = tempRating
            #Updates the attacker ratings per operator
            if d >= 34:
                tempAtkRating.append(Op_array[matchingValue][9+12*d])
                
                
                tempAtkRounds.append(Op_array[matchingValue][20+12*d])
            else:
                
                tempDefRating.append(Op_array[matchingValue][9+12*d])
                tempDefRounds.append(Op_array[matchingValue][20+12*d])
        #Puts the max ratings and rounds into the array
        maxAtkRating = max(tempAtkRating)
        maxDefRating = max(tempDefRating)
        maxAtkRounds = max(tempAtkRounds)
        maxDefRounds = max(tempDefRounds)
        Op_array[matchingValue][1] = Operators.OperatorsValues[34+tempAtkRounds.index(maxAtkRounds)]
        Op_array[matchingValue][2] = maxAtkRounds
        Op_array[matchingValue][3] = Operators.OperatorsValues[tempDefRounds.index(maxDefRounds)]
        Op_array[matchingValue][4] = maxDefRounds
        Op_array[matchingValue][5] = Operators.OperatorsValues[34+tempAtkRating.index(maxAtkRating)]
        Op_array[matchingValue][6] = maxAtkRating
        Op_array[matchingValue][7] = Operators.OperatorsValues[tempDefRating.index(maxDefRating)]
        Op_array[matchingValue][8] = maxDefRating
            
            



#Check new or returning players for the Stats Excel Page
for column_index, value in enumerate(OperatorHeader, start=1):
    cell = sheet1.cell(row=1, column=column_index)
    cell.value = value

for row_index, row in enumerate(Op_array, start=2):
    for column_index, value in enumerate(row, start=1):
        cell = sheet1.cell(row=row_index, column=column_index)
        cell.value = value
#Checks the users to see if they already match
for i in range(len(Users)):
    matching = 0
    matchingValue = -1
    for j in range(len(ExcelUsername)):
        if ExcelUsername[j] == Users[i]:
            matching = matching + 1
            matchingValue = j
    #If do not match append to the arrays
    if matching == 0:
        ExcelUsername.append(Users[i])
        ExcelRating.append(round(ratingSys(Kills[i], Kills[i]/Deaths[i], MK[i], EKills[i] - EDeaths[i], Plant[i], Clutch[i], KOST[i]/Rounds[i], (Rounds[i]-Deaths[i])/Rounds[i], Rounds[i]),2))
        ExcelKills.append(Kills[i])
        ExcelDeaths.append(Deaths[i])
        ExcelKD.append(round(Kills[i]/Deaths[i], 2))
        ExcelEK.append(EKills[i])
        ExcelED.append(EDeaths[i])
        ExcelEntry.append(EKills[i] - EDeaths[i])
        ExcelKOST.append(round(KOST[i]/Rounds[i],2))
        ExcelKPR.append(round(Kills[i]/Rounds[i],2))
        ExcelSRV.append(round(((Rounds[i]-Deaths[i])/Rounds[i]),2))
        ExcelMKills.append(round(MK[i]/Rounds[i],2))
        ExcelTrade.append(round(Trade[i]/Deaths[i],2))
        ExcelClutch.append(round(Clutch[i]/Rounds[i],2))
        ExcelPlants.append(round(Plant[i]/(Rounds[i]/2),2))
        ExcelHS.append(math.ceil(HS[i]))
        ExcelRound.append(Rounds[i])
    #If match, update the value and replace at position
    else:
        ExcelHS[matchingValue] = math.ceil((ExcelHS[matchingValue]*ExcelKills[matchingValue] + HS[i]*Kills[i])/(ExcelKills[matchingValue] + Kills[i]))
        ExcelKills[matchingValue] = ExcelKills[matchingValue] + Kills[i]
        ExcelDeaths[matchingValue] = ExcelDeaths[matchingValue] + Deaths[i]
        ExcelKD[matchingValue] = round(ExcelKills[matchingValue]/ExcelDeaths[matchingValue],2)
        ExcelEK[matchingValue] = ExcelEK[matchingValue] + EKills[i]
        ExcelED[matchingValue] = ExcelED[matchingValue] + EDeaths[i]
        ExcelEntry[matchingValue] = ExcelEK[matchingValue] - ExcelED[matchingValue]
        ExcelKOST[matchingValue] = round((ExcelKOST[matchingValue]*ExcelRound[matchingValue] + KOST[i])/(ExcelRound[matchingValue] + Rounds[i]),2)
        ExcelKPR[matchingValue] = round((ExcelKPR[matchingValue]*ExcelRound[matchingValue] + Kills[i])/(ExcelRound[matchingValue] + Rounds[i]),2)
        ExcelSRV[matchingValue] = round(((ExcelRound[matchingValue] + Rounds[i]) - ExcelDeaths[matchingValue])/(ExcelRound[matchingValue] + Rounds[i]),2)
        ExcelMKills[matchingValue] = round((ExcelMKills[matchingValue]*ExcelRound[matchingValue] + MK[i])/(ExcelRound[matchingValue]+Rounds[i]),2)
        ExcelTrade[matchingValue] = round((ExcelTrade[matchingValue]*ExcelRound[matchingValue] + Trade[i])/(ExcelRound[matchingValue]+Rounds[i]),2)
        ExcelClutch[matchingValue] = round((ExcelClutch[matchingValue]*ExcelRound[matchingValue] + Clutch[i])/(ExcelRound[matchingValue]+Rounds[i]),2)
        ExcelPlants[matchingValue] = round((ExcelPlants[matchingValue]*ExcelRound[matchingValue] + Plant[i])/(ExcelRound[matchingValue]+Rounds[i]),2)
        
        ExcelRound[matchingValue] = ExcelRound[matchingValue] + Rounds[i]
        ExcelRating[matchingValue] = round(ratingSys(ExcelKills[matchingValue], ExcelKD[matchingValue], ExcelMKills[matchingValue], ExcelEntry[matchingValue], ExcelPlants[matchingValue], ExcelClutch[matchingValue], ExcelKOST[matchingValue], ExcelSRV[matchingValue], ExcelRound[matchingValue]),2)
 



            



#Reput in the cell values now that they are updated
for i in range(len(ExcelUsername)):
    cell = sheet[ExcelCols[0] + str(i+2)]
    cell.value = ExcelUsername[i]

for i in range(len(ExcelRating)):
    cell = sheet[ExcelCols[1] + str(i+2)]
    cell.value = ExcelRating[i]

for i in range(len(ExcelKills)):
    cell = sheet[ExcelCols[2] + str(i+2)]
    cell.value = ExcelKills[i]

for i in range(len(ExcelDeaths)):
    cell = sheet[ExcelCols[3] + str(i+2)]
    cell.value = ExcelDeaths[i]

for i in range(len(ExcelKD)):
    cell = sheet[ExcelCols[4] + str(i+2)]
    cell.value = ExcelKD[i]

for i in range(len(ExcelEK)):
    cell = sheet[ExcelCols[5] + str(i+2)]
    cell.value = ExcelEK[i]

for i in range(len(ExcelED)):
    cell = sheet[ExcelCols[6] + str(i+2)]
    cell.value = ExcelED[i]

for i in range(len(ExcelEntry)):
    cell = sheet[ExcelCols[7] + str(i+2)]
    cell.value = ExcelEntry[i]

for i in range(len(ExcelKOST)):
    cell = sheet[ExcelCols[8] + str(i+2)]
    cell.value = ExcelKOST[i]

for i in range(len(ExcelKPR)):
    cell = sheet[ExcelCols[9] + str(i+2)]
    cell.value = ExcelKPR[i]

for i in range(len(ExcelSRV)):
    cell = sheet[ExcelCols[10] + str(i+2)]
    cell.value = ExcelSRV[i]

for i in range(len(ExcelMKills)):
    cell = sheet[ExcelCols[11] + str(i+2)]
    cell.value = ExcelMKills[i]

for i in range(len(ExcelTrade)):
    cell = sheet[ExcelCols[12] + str(i+2)]
    cell.value = ExcelTrade[i]

for i in range(len(ExcelClutch)):
    cell = sheet[ExcelCols[13] + str(i+2)]
    cell.value = ExcelClutch[i]

for i in range(len(ExcelPlants)):
    cell = sheet[ExcelCols[14] + str(i+2)]
    cell.value = ExcelPlants[i]

for i in range(len(ExcelHS)):
    cell = sheet[ExcelCols[15] + str(i+2)]
    cell.value = ExcelHS[i]

for i in range(len(ExcelRound)):
    cell = sheet[ExcelCols[16] + str(i+2)]
    cell.value = ExcelRound[i]

for h in range(len(Users)):
    temp = [Users[h]]

# Save the changes
#This is how to get the Excel file to save the changes that were made
workbook.save(filename=Excel)


#These are examples of how to access the data in python per operator stats
#players[9].SingleOperatorStats('Flores')
#players[7].AllOps()








