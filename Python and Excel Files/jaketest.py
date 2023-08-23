import json
import math
import numpy as np
import statistics
import openpyxl
from openpyxl import load_workbook


operatorsToNum = {"Mute": 0, "Smoke": 1, "Castle": 2, "Pulse": 3, "Doc": 4, "Rook": 5, "Jager": 6, "Bandit": 7,"Tachanka": 8, "Kapkan": 9, "Frost": 10, "Valkyrie": 11, "Caveira": 12, "Echo": 13, "Mira": 14,"Lesion": 15, "Ela": 16, "Vigil": 17, "Alibi": 18, "Maestro": 19, "Clash": 20, "Kaid": 21, "Mozzie": 22,"Warden": 23, "Goyo": 24, "Wamai": 25, "Oryx": 26, "Melusi": 27, "Aruni": 28, "Thunderbird": 29,"Thorn": 30, "Azami": 31, "Solis": 32, "Fenrir": 33, "Sledge": 34, "Thatcher": 35, "Ash": 36,"Thermite": 37, "Montagne": 38, "Twitch": 39, "Blitz": 40, "IQ": 41, "Fuze": 42, "Glaz": 43,"Buck": 44, "Blackbeard": 45, "Capitao": 46, "Hibana": 47, "Jackal": 48, "Ying": 49, "Zofia": 50,"Dokkaebi": 51, "Finka": 52, "Lion": 53, "Maverick": 54, "Nomad": 55, "Gridlock": 56, "Nokk": 57,"Amaru": 58, "Kali": 59, "Iana": 60, "Ace": 61, "Zero": 62, "Flores": 63, "Osa": 64, "Sens": 65,"Grim": 66, "Brava": 67}
numToOperators = {0: 'Mute', 1: 'Smoke', 2: 'Castle', 3: 'Pulse', 4: 'Doc', 5: 'Rook', 6: 'Jager', 7: 'Bandit', 8: 'Tachanka', 9: 'Kapkan', 10: 'Frost', 11: 'Valkyrie', 12: 'Caveira', 13: 'Echo', 14: 'Mira', 15: 'Lesion', 16: 'Ela', 17: 'Vigil', 18: 'Alibi', 19: 'Maestro', 20: 'Clash', 21: 'Kaid', 22: 'Mozzie', 23: 'Warden', 24: 'Goyo', 25: 'Wamai', 26: 'Oryx', 27: 'Melusi', 28: 'Aruni', 29: 'Thunderbird', 30: 'Thorn', 31: 'Azami', 32: 'Solis', 33: 'Fenrir', 34: 'Sledge', 35: 'Thatcher', 36: 'Ash', 37: 'Thermite', 38: 'Montagne', 39: 'Twitch', 40: 'Blitz', 41: 'IQ', 42: 'Fuze', 43: 'Glaz', 44: 'Buck', 45: 'Blackbeard', 46: 'Capitao', 47: 'Hibana', 48: 'Jackal', 49: 'Ying', 50: 'Zofia', 51: 'Dokkaebi', 52: 'Finka', 53: 'Lion', 54: 'Maverick', 55: 'Nomad', 56: 'Gridlock', 57: 'Nokk', 58: 'Amaru', 59: 'Kali', 60: 'Iana', 61: 'Ace', 62: 'Zero', 63: 'Flores', 64: 'Osa', 65: 'Sens', 66: 'Grim', 67: 'Brava'}
#File Path Location of Excel Sheet
excelFile = 'C:\\Users\\JXG3061\\Desktop\\Jake\\Code Testing\\JakeStats.xlsx'
#excelFile = 'C:\\Users\\jakeg\\OneDrive\\Desktop\\r6-dissect-v0.11.1-windows-amd64\\Stats.xlsx'
#Load the workbook into read cells
workbook = load_workbook(filename=excelFile)
timeToTrade = 10 #seconds
opNumbers = 68 #number of operators
# <= 33 is defender, > 33 is attacker
#Header of 'Operator Stats' page in Excel
operatorHeaderInfo = [' Rating', ' Kills', ' Deaths', ' Entry Kills', ' Entry Deaths', ' KOST', ' HS', ' MultiKills', ' Trades', ' Clutch', ' Plants', ' Rounds']
operatorHeader = []
operatorHeader.append('Username')
operatorHeader.append('Most Played Atk Op')
operatorHeader.append('# of Rounds')
operatorHeader.append('Most Played Def Op')
operatorHeader.append('# of Rounds')
operatorHeader.append('Highest Rating Atk Op')
operatorHeader.append('Rating')
operatorHeader.append('Highest Rating Def Op')
operatorHeader.append('Rating')
for i in range(opNumbers):
    for j in range(len(operatorHeaderInfo)):
        tempStr = numToOperators[i] + operatorHeaderInfo[j]
        operatorHeader.append(tempStr)


# Select the desired sheet
sheetName = 'Stats'
opSheetName = 'Operator Stats'
excelMainSheet = workbook[sheetName]
excelOpStatSheet = workbook[opSheetName]
#Define variables
excelUsername = []
excelRating = []
excelKills = []
excelDeaths = []
excelKD = []
excelEntryKill = []
excelEntryDeath = []
excelEntryPlusMinus = []
excelKOST = []
excelKPR = []
excelSRV = []
excelMKills = []
excelTrade = []
excelClutch = []
excelPlants = []
excelHS = []
excelAtk = []
excelDef = []
excelRound = []
excelColUsername = 'A'
excelColRating = 'B'
excelColKills = 'C' 
excelColDeaths = 'D'
excelColKD = 'E'
excelColEK = 'F'
excelColED = 'G'
excelColEntry = 'H'
excelColKOST = 'I'
excelColKPR = 'J'
excelColSRV = 'K'
excelColMKills = 'L'
excelColTrade = 'M'
excelColClutch = 'N'
excelColPlants = 'O'
excelColHS = 'P'
excelColRound = 'Q'

#Array of the column names and arrays that will hold data to make it easier to access individual columns below
excelCols = [excelColUsername,excelColRating, excelColKills,excelColDeaths,excelColKD,excelColEK,excelColED,excelColEntry,excelColKOST,excelColKPR,excelColSRV,excelColMKills,excelColTrade,excelColClutch,excelColPlants,excelColHS,excelColRound]
excelArr = [excelUsername, excelRating, excelKills, excelDeaths, excelKD, excelEntryKill, excelEntryDeath, excelEntryPlusMinus, excelKOST, excelKPR, excelSRV, excelMKills, excelTrade, excelClutch, excelPlants, excelHS, excelRound]

# Loops through all excel columns to store data, removes the first element bc this is the header 
for i in range(len(excelCols)):
    for cell in excelMainSheet[excelCols[i]]:
        excelArr[i].append(cell.value)
    excelArr[i].pop(0)



#Takes in the length of the ExcelUsername for 'Operator Stats' to read that many lines from the Operator Stats and read down the row
#Creates a 2d array of width of array of ExcelUsername and Length of # of players
opArray = []

if len(excelUsername) != 0:
    for i in range(len(excelUsername)):
        tempArr = []
        for cell in excelOpStatSheet[2+i]:
            tempArr.append(cell.value)
        opArray.append(tempArr)












#Definiton of rating system, same as used in basic stats, !!SHOULD BECOME UNIFORM!!
def ratingSys(rKills, rKD, rMK, rEntry, rPlants, rClutch, rKOST, rSRV, rRounds):
    rating = (rKD**2 + 0.4*(rKills) + 0.15*rMK)/rRounds + 0.75*(rEntry)/rRounds + (rPlants + rClutch)/rRounds + rKOST + rSRV/3
    return rating

#Prints the header separate from the match files now
def headerPrint(mapPlayed, team1Score, team2Score, team1Name, team2Name):
        print('Map: ' + mapPlayed)
        print(team1Name + ' - ' + team2Name)
        print(str(team1Score) + ' - ' + str(team2Score))
        print('')
#Creates Structs for each player and their stats
class basicStats:
    #Initializes variables, (Why do I have to do this?)
    def __init__(self, username, kills, deaths, kD, eKills, eDeaths, entry, kOST, kPR, sRV, mKills, trade, clutch, plants, defuse, hsPercent, favAtk, favDef, rounds, opK, opD, opKD, opEK, opED, opEntry, opKOST, opKPR, opSRV, opMKills, opTrade, opClutch, opPlants, opDefuse, opHS, opRounds, mapPlayed, team1Score, team2Score, team1Name, team2Name):
        self.username = username
        self.kills = kills
        self.deaths = deaths
        self.kD = kD
        self.eKills = eKills
        self.eDeaths = eDeaths
        self.entry = entry
        self.kOST = kOST
        self.kPR = kPR
        self.sRV = sRV
        self.mKills = mKills
        self.trade = trade
        self.clutch = clutch
        self.plants = plants
        self.defuse = defuse
        self.hsPercent = hsPercent
        self.favAtk = favAtk
        self.favDef = favDef
        self.rounds = rounds
        self.opK = opK
        self.opD = opD
        self.opKD = opKD
        self.opEK = opEK
        self.opED = opED
        self.opEntry = opEntry
        self.opKOST = opKOST
        self.opKPR = opKPR
        self.opSRV = opSRV
        self.opMKills = opMKills
        self.opTrade = opTrade
        self.opClutch = opClutch
        self.opPlants = opPlants
        self.opDefuse = opDefuse
        self.opHS = opHS
        self.opRounds = opRounds
        self.mapPlayed = mapPlayed
        self.team1Score = team1Score
        self.team2Score = team2Score
        self.team1Name = team1Name
        self.team2Name = team2Name
        
#Rating System, very basic
#Same rating system as before
    # def rating(self):
    #     rating = (self.KD**2 + 0.4*(self.Kills) + 0.15*self.MKills)/self.Rounds + 0.75*(self.Entry)/self.Rounds + (self.Plants + self.Clutch)/self.Rounds + self.KOST + self.SRV/3
    #     return rating
#Prints out all 'relevant' stats, in similar format as siegeGG, adds multikills and trades for greater visibility
    
    def printIndivStat(self, intro):
        #defines the rating in this def
        rating = ratingSys(self.kills, self.kD, self.mKills, self.entry, self.plants, self.clutch, self.kOST, self.sRV, self.rounds)
        #if the first user printed, prints a header
        
        #Top part of the print, gives the map, and score and header
        if(intro == 1):
            
            formattedString = "{:<15} | {:<6} | {:<10} | {:<8} | {:<6} | {:<6} | {:<6} | {:<6} | {:<6} | {:<6} | {:<6} | {:<6} | {:<10} | {:<10}".format(
                'Username', 'Rating', 'K-D(KD)', 'Entry', 'KOST', 'KPR', 'SRV', '1vX', 'Plants', 'Multis', 'Trades', 'HS', 'Attacker', 'Defender', '', ''
            )
            #copies the underlined text to the length of the text above and creates a line
            underlinedString = formattedString + '\n' + '-' * len(formattedString)
            print(underlinedString)
        
            

        #formats the plus or minus in front of the KD and entry to make it + or -
        plusMinus = self.kills - self.deaths
        if(plusMinus > 0):
            strKD = str(self.kills) + '-' + str(self.deaths) + '(+' + str(plusMinus)+')'
        else:
            strKD = str(self.kills) + '-' + str(self.deaths) + '(' + str(plusMinus)+')'
        
        #Same formatting for Entry Stats
        ePlusMinus = self.eKills - self.eDeaths
        if(ePlusMinus > 0):
            strEntry = str(int(self.eKills)) + '-' + str(int(self.eDeaths)) + '(+'+str(int(ePlusMinus))+')'
        else:
            strEntry = str(int(self.eKills)) + '-' + str(int(self.eDeaths)) + '('+str(int(ePlusMinus))+')'
        #formatting the text for each user and prints
        formatKOST = "{:.2f}".format(self.kOST)
        formatKPR = "{:.2f}".format(self.kPR)
        formatRating = "{:.2f}".format(rating)
        formattedString = "{:<15} | {:<6} | {:<10} | {:<8} | {:<6} | {:<6} | {:<6} | {:<6} | {:<6} | {:<6} | {:<6} | {:<6} | {:<10} | {:<10}".format(
            self.username,
            formatRating,
            strKD,
            strEntry,
            formatKOST,
            formatKPR,
            str(int(self.sRV*100)) + '%',
            int(self.clutch),
            int(self.plants),
            int(self.mKills),
            int(self.trade),
            str(int(self.hsPercent)) + '%',
            self.favAtk,
            self.favDef
        )
        print(formattedString)
    #Define ratings for Operator Ratings for a player
    def operatorRating(self):
        operatorRating = np.zeros(opNumbers)
        #rating = (1.5*self.KD + 0.25*(self.Kills) + 0.15*self.MKills)/self.Rounds + 0.75*(self.Entry)/self.Rounds + (self.Plants + self.Clutch)/self.Rounds + self.KOST + self.SRV/3
        for j in (range(opNumbers)):
            if self.opRounds[j] == 0:
                operatorRating[j] = 0
            else:
                operatorRating[j] = ratingSys(self.opK[j], self.opKD[j], self.opMKills[j], self.opEntry[j], self.opPlants[j], self.opClutch[j], self.opKOST[j], self.opSRV[j], self.opRounds[j])
        return operatorRating
    #Simple Way to read all player Operator Rating, just in python currently, but could be phased out
    def allOps(self):
        Op = self.operatorRating()
        for k in (range(opNumbers)):
            number_str = Op[k]
            roundedRating = "{:.2f}".format(float(number_str))
            print(numToOperators[k] + ': ' + roundedRating)
    #Similar to above, can look at individual player and full rating for a single operator, python only, either phase out or can be used in maybe a different aspect
    #Very similar formatting to the full list for a single map, should be merged
    def singleOperatorStats(self, inputStr):
        operatorValue = operatorsToNum[inputStr]
        print('\n')
        opsRate = self.operatorRating()
        
        formattedString = "{:<12} | {:<10} | {:<6} | {:<10} | {:<8} | {:<6} | {:<6} | {:<6} | {:<6} | {:<6} | {:<6} | {:<6} | {:<6} | {:<6}".format(
                'Player','Op Name', 'Rating', 'K-D(KD)', 'Entry', 'KOST', 'KPR', 'SRV', '1vX', 'Plants', 'Multis', 'Trades', 'HS', 'Rounds', '')
        underlinedString = formattedString + '\n' + '-' * len(formattedString)
        print(underlinedString)
        plusMinus = self.opK[operatorValue] - self.opD[operatorValue]
        if(plusMinus > 0):
            strKD = str(int(self.opK[operatorValue])) + '-' + str(int(self.opD[operatorValue])) + '(+' + str(int(plusMinus))+')'
        else:
            strKD = str(int(self.opK[operatorValue])) + '-' + str(int(self.opD[operatorValue])) + '(' + str(int(plusMinus))+')'
        
        ePlusMinus = self.opEK[operatorValue] - self.opED[operatorValue]
        if(ePlusMinus > 0):
            strEntry = str(int(self.opEK[operatorValue])) + '-' + str(int(self.opED[operatorValue])) + '(+'+str(int(ePlusMinus))+')'
        else:
            strEntry = str(int(self.opEK[operatorValue])) + '-' + str(int(self.opED[operatorValue])) + '('+str(int(ePlusMinus))+')'
        #formatting the text for each user and prints
        if self.opK[operatorValue] == 0:
            hs = 0
        else:
            hs = self.opHS[operatorValue]/self.opK[operatorValue]*100
        formatKOST = "{:.2f}".format(self.opKOST[operatorValue])
        formatKPR = "{:.2f}".format(self.opKPR[operatorValue])
        formatRating = "{:.2f}".format(opsRate[operatorValue])
        
        formattedString = "{:<12} | {:<10} | {:<6} | {:<10} | {:<8} | {:<6} | {:<6} | {:<6} | {:<6} | {:<6} | {:<6} | {:<6} | {:<6} | {:<6}".format(
            self.username,
            inputStr,
            formatRating,
            strKD,
            strEntry,
            formatKOST,
            formatKPR,
            str(int(self.opSRV[operatorValue]*100)) + '%',
            int(self.opClutch[operatorValue]),
            int(self.opPlants[operatorValue]),
            int(self.opMKills[operatorValue]),
            int(self.opTrade[operatorValue]),
            str(int(hs)) + '%',
            str(int(self.opRounds[operatorValue]))
        )
        print(formattedString)
            



#Parses json file for all stats
#Unable to determine how the disable works !!NEEDS TESTING!!
def singleMap(dict):
    #define all variables to pass through/use
    team1Score = 0
    team2Score = 0
    #Takes the team names from the first round
    team1Name = dict["rounds"][0]["teams"][0]["name"]
    team2Name = dict["rounds"][0]["teams"][1]["name"]
    usernameList = []
    usernameLookup = {}
    killAmount = []
    deathAmount = []
    hsPercent = []
    roundCount = []
    atkMain = []
    defMain = []
    #Note: The operator arrays are filled to -1, this is because if it was initialized at 0, could show a false positive of mute being played because he is operator 0
    roundOps = np.full(10,-1)
    dTotalOps = np.array([])
    aTotalOps = np.array([])
    aRoundOps = np.full(10,-1)
    dRoundOps = np.full(10,-1)
    aMainOp = np.array([])
    dMainOp = np.array([])
    winningTeamMembers = []
    entryKills = np.zeros(10)
    entryDeaths = np.zeros(10)
    plants = np.zeros(10)
    defusal = np.zeros(10)
    kOSTRounds = np.zeros(10)
    kOSTTotal=np.zeros(10)
    kOSTSurv = np.zeros(10)
    clutches = np.zeros(10)
    multikills = np.zeros(10)
    trades = np.zeros(10)
    opKills = np.zeros((10,68))
    opDeaths = np.zeros((10,68))
    opEKills = np.zeros((10,68))
    opEDeaths = np.zeros((10,68))
    opHS = np.zeros((10,68))
    opPlants = np.zeros((10,68))
    opDefusal = np.zeros((10,68))
    opKOST = np.zeros((10,68))
    opKOSTRound = np.zeros((10,68))
    opClutches = np.zeros((10,68))
    opMultikills = np.zeros((10,68))
    opTrades = np.zeros((10,68))
    opRounds = np.zeros((10,68))
    
    #Takes the first round map name to output
    map = dict["rounds"][0]["map"]["name"]
    #takes the total number of kills in the map and hs percent and rounds played
    for i in range(10):
        #Appends list at the end of rounds to track the basic stats that is given, could be improved if needed
        usernameList.append(dict["stats"][i]["username"])
        killAmount.append(dict["stats"][i]["kills"])
        deathAmount.append(dict["stats"][i]["deaths"])
        hsPercent.append(dict["stats"][i]["headshotPercentage"])
        roundCount.append(dict["stats"][i]["rounds"])
        #creates a lookup table for each player, 0-9 based on name and how incrememented in the json file
        #Makes it easier to upload names to further stats, because can organize players
        usernameLookup[usernameList[i]] = i
    #large loop to look at all rounds
    #!! Restructing Note, should do a double for loop where outer loop is rounds, and inner loops in actions in rounds !!
    for i in range(len(dict["rounds"])):
        actions = len(dict["rounds"][i]["matchFeedback"])
        
        #actions is number of occurences in the 'main phase' of each round, contains everything that a player can do to impact each round
        #!!2 Different logics between while and for loop, should restructure!!
        
        #Updates team score each round, so the final score is output at the end of all rounds
        team1Score = dict["rounds"][i]["teams"][0]["score"]
        team2Score = dict["rounds"][i]["teams"][1]["score"]
        actions = len(dict["rounds"][i]["matchFeedback"])
        #loops through all the actions looking for the first kill to occur
        aRoundOps = np.full(10,-1)
        dRoundOps = np.full(10,-1)
        if(dict["rounds"][i]["teams"][0]["won"] == True):
            WinningTeam = 0
        else:
            WinningTeam = 1
        winningTeamMembers = []
        clutchPlayer = ''
        clutchAlive = 5
        for v in range(len(dict["rounds"][i]["players"])):
            #tracks clutching based on if your team won and you were the only player on your team alive
            
            if(dict["rounds"][i]["players"][v]["teamIndex"] == WinningTeam):
                winningTeamMembers.append(dict["rounds"][i]["players"][v]["username"])
            #start with 5 players alive on the winning team, if a winning team member dies, then reduce the number alive
            #add winning round members
            
            
            #if the operator that is selected by the player is > 33 it is an attacker by my dictionary, otherwise its a defender
            #if an attacker, needs to check if a repick occurs
            if(operatorsToNum[dict["rounds"][i]["players"][v]["operator"]["name"]] > 33):
                #Puts the Player on the operator that they played that round in their 'spot' in the array, and given a numerical value from the dict 'Operators'
                roundOps[usernameLookup[dict["rounds"][i]["players"][v]["username"]]] = operatorsToNum[dict["rounds"][i]["players"][v]["operator"]["name"]]
                aRoundOps[usernameLookup[dict["rounds"][i]["players"][v]["username"]]] = operatorsToNum[dict["rounds"][i]["players"][v]["operator"]["name"]]
                #as attackers can swap in prep phase, look for the match feedback for an operator swap and update the value
                #Checks through all the round actions to see if an operator is swapped off
                for c in range(actions):
                    if(dict["rounds"][i]["matchFeedback"][c]["type"]["name"] == "OperatorSwap"):
                        #Updates the value in similar value
                        roundOps[usernameLookup[dict["rounds"][i]["matchFeedback"][c]["username"]]] = operatorsToNum[dict["rounds"][i]["matchFeedback"][c]["operator"]["name"]]
                        aRoundOps[usernameLookup[dict["rounds"][i]["matchFeedback"][c]["username"]]] = operatorsToNum[dict["rounds"][i]["matchFeedback"][c]["operator"]["name"]]
                        
            else:
                #Defenders do not change, so the first time seen can be set
                roundOps[usernameLookup[dict["rounds"][i]["players"][v]["username"]]] = operatorsToNum[dict["rounds"][i]["players"][v]["operator"]["name"]]
                dRoundOps[usernameLookup[dict["rounds"][i]["players"][v]["username"]]] = operatorsToNum[dict["rounds"][i]["players"][v]["operator"]["name"]]
            #Hard to read
            #Puts all the operator plays into the 2d array for the Operator Stats Page, Puts the username lookup and and operator location and increments the play count
            
            opRounds[usernameLookup[dict["rounds"][i]["players"][v]["username"]],roundOps[usernameLookup[dict["rounds"][i]["players"][v]["username"]]]] = opRounds[usernameLookup[dict["rounds"][i]["players"][v]["username"]],roundOps[usernameLookup[dict["rounds"][i]["players"][v]["username"]]]] + 1
        #use the dict defined numerical value of the op to add to a total array
        aTotalOps = np.append(aTotalOps, aRoundOps)
        dTotalOps = np.append(dTotalOps, dRoundOps)
        #Loops for opening kill
        firstKill = 0
        kOSTRounds = np.zeros(10)
        #If did nothing but lived, should have 1
        kOSTSurv = np.ones(10)
        opKOSTRound = np.zeros((10,opNumbers))
        for j in range(actions):
            
            if firstKill == 0 and dict["rounds"][i]["matchFeedback"][j]["type"]["name"] == "Kill":
                firstKill = firstKill + 1
                entryKills[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["username"]]] = entryKills[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["username"]]] + 1
                entryDeaths[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["target"]]] = entryDeaths[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["target"]]] + 1
                opEKills[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["username"]],roundOps[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["username"]]]] = opEKills[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["username"]],roundOps[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["username"]]]] + 1
                opEDeaths[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["target"]],roundOps[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["target"]]]] = opEDeaths[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["target"]],roundOps[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["target"]]]] + 1
            #gives the planter a point if they get defuser down and what operator they played
            if dict["rounds"][i]["matchFeedback"][j]["type"]["name"] == "DefuserPlantComplete":
                plants[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["username"]]] = plants[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["username"]]] + 1
                opPlants[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["username"]],roundOps[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["username"]]]] = opPlants[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["username"]],roundOps[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["username"]]]]+1
            if(dict["rounds"][i]["matchFeedback"][j]["type"]["name"] == "Kill"):
                #If a kill happens, the round is counted for KOST, but not survival
                kOSTRounds[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["username"]]] = 1
                kOSTSurv[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["target"]]] = 0
                #looks for trades, when a user dies, capture tOD and who killed them
                #then iterate through until the time is past the time to trade to see if the trade occurs
                timeOfDeath = dict["rounds"][i]["matchFeedback"][j]["timeInSeconds"]
                userToBeTraded = dict["rounds"][i]["matchFeedback"][j]["username"]
                l=0
                #If the time happens, loops through all actions to see if it is within the 10 second trade time, otherwise go to next action
                while(timeOfDeath-timeToTrade < dict["rounds"][i]["matchFeedback"][j+l]["timeInSeconds"]):
                    if(dict["rounds"][i]["matchFeedback"][j+l]["type"]["name"] == "Kill" and dict["rounds"][i]["matchFeedback"][j+l]["target"] == userToBeTraded):
                        kOSTRounds[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["target"]]] = 1
                        trades[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["target"]]] = trades[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["target"]]] + 1
                        opTrades[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["target"]],roundOps[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["target"]]]] = opTrades[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["target"]],roundOps[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["target"]]]] + 1
                        break
                    else:
                        l = l+1
                        if(j+l == actions):
                            break
                
            #adds plant if get plant down
            elif(dict["rounds"][i]["matchFeedback"][j]["type"]["name"] == "DefuserPlantComplete"):
                kOSTRounds[usernameLookup[dict["rounds"][i]["matchFeedback"][j]["username"]]] = 1
        
        
            

        #tracks KOST (Kills, Objectives[plants], Survival, Trades)
        
        #looks for survival rate and trades
        
            
        #adds a round where you add to your kost, makes sure not to duplicate if 2 things occur in a round
        #Adds rounds to if they survived
        kOSTRounds = kOSTRounds + kOSTSurv
        
        #If higher than one, sets back to 1 to not overrate the KOST rounds if someone does more than one KOST action in a round
        for n in range(len(kOSTRounds)):
            if(kOSTRounds[n] > 1):
                kOSTRounds[n] = 1
            opKOSTRound[n, roundOps[n]] = kOSTRounds[n]
        #Adds the KOST to total number of rounds that players achieved during the number of rounds
        kOSTTotal = kOSTTotal + kOSTRounds
        opKOST = opKOST + opKOSTRound

        #tracks each user for their multikills over a game
        #Checks for HS percentage and Kills for operators per round, also looks to see if number of kills was greater than 1
        for g in range(len(dict["rounds"][i]["stats"])):
            
            opKills[usernameLookup[dict["rounds"][i]["stats"][g]["username"]],roundOps[usernameLookup[dict["rounds"][i]["stats"][g]["username"]]]] = opKills[usernameLookup[dict["rounds"][i]["stats"][g]["username"]],roundOps[usernameLookup[dict["rounds"][i]["stats"][g]["username"]]]] + dict["rounds"][i]["stats"][g]["kills"]
            opHS[usernameLookup[dict["rounds"][i]["stats"][g]["username"]],roundOps[usernameLookup[dict["rounds"][i]["stats"][g]["username"]]]] = opHS[usernameLookup[dict["rounds"][i]["stats"][g]["username"]],roundOps[usernameLookup[dict["rounds"][i]["stats"][g]["username"]]]] + dict["rounds"][i]["stats"][g]["headshots"]
            if(dict["rounds"][i]["stats"][g]["died"] == True):
                opDeaths[usernameLookup[dict["rounds"][i]["stats"][g]["username"]],roundOps[usernameLookup[dict["rounds"][i]["stats"][g]["username"]]]] = opDeaths[usernameLookup[dict["rounds"][i]["stats"][g]["username"]],roundOps[usernameLookup[dict["rounds"][i]["stats"][g]["username"]]]] + 1
            if(dict["rounds"][i]["stats"][g]["kills"] > 1):
                multikills[usernameLookup[dict["rounds"][i]["stats"][g]["username"]]] = multikills[usernameLookup[dict["rounds"][i]["stats"][g]["username"]]] + 1
                opMultikills[usernameLookup[dict["rounds"][i]["stats"][g]["username"]], roundOps[usernameLookup[dict["rounds"][i]["stats"][g]["username"]]]] = opMultikills[usernameLookup[dict["rounds"][i]["stats"][g]["username"]], roundOps[usernameLookup[dict["rounds"][i]["stats"][g]["username"]]]] + 1
        
        
            
        
        #json file has the stat added if a clutch occurs of 1vX, so if it exists then a clutch occured
        for a in range(len(dict["rounds"][i]["stats"])):
            if "1vX" in dict["rounds"][i]["stats"][a]:
                clutches[usernameLookup[dict["rounds"][i]["stats"][a]["username"]]] = clutches[usernameLookup[dict["rounds"][i]["stats"][a]["username"]]] + 1
                opClutches[usernameLookup[dict["rounds"][i]["stats"][a]["username"]],roundOps[usernameLookup[dict["rounds"][i]["stats"][a]["username"]]]] = opClutches[usernameLookup[dict["rounds"][i]["stats"][a]["username"]],roundOps[usernameLookup[dict["rounds"][i]["stats"][a]["username"]]]] + 1
        #Check again to see that only 1 player was alive on the winning team to confirm that it is a clutch and give it to player and the operator they played that round
        
            

        
        
        #find all main ops by using array math to loop through and find each players most played operator
    
    for z in range(len(dict["rounds"][i]["players"])):
        aMainOp = np.array([])
        dMainOp = np.array([])
        #Separate players into their individual operators to check for main operator that a single person
        for q in range(len(dict["rounds"])):
            aMainOp = np.append(aMainOp, aTotalOps[q*10 + z])
            aMainOp = aMainOp[aMainOp>=0]
            dMainOp = np.append(dMainOp, dTotalOps[q*10 + z])
            dMainOp = dMainOp[dMainOp>=0]
        #add each users main op back to main array
        #Create an array to check what the main operator each person played
        aMain = statistics.mode(aMainOp)
        aMain = int(aMain)
        dMain = statistics.mode(dMainOp)
        atkMain.append(numToOperators[aMain])
        defMain.append(numToOperators[dMain])
     
        
    #all return values are array values of # of players
    return [usernameList,killAmount,deathAmount,entryKills,entryDeaths,kOSTTotal, hsPercent,multikills,trades, clutches, plants,defusal,atkMain,defMain,roundCount, opKills, opDeaths, opEKills, opEDeaths, opKOST, opHS, opMultikills, opTrades, opClutches, opPlants,opDefusal, opRounds, map, team1Score, team2Score, team1Name, team2Name]
        


with open("C:\\Users\\JXG3061\\Desktop\\Jake\\Code Testing\\scrim1.json", 'r') as f:
    my_dict = json.load(f)
with open("C:\\Users\\JXG3061\\Desktop\\Jake\\Code Testing\\scrim2.json", 'r') as f:
    my_dict1 = json.load(f)
#open the json file i am parsing
# with open("C:\\Users\\jakeg\\OneDrive\\Desktop\\r6-dissect-v0.11.1-windows-amd64\\scrim1.json", 'r') as f:
#     my_dict = json.load(f)
# with open("C:\\Users\\jakeg\\OneDrive\\Desktop\\r6-dissect-v0.11.1-windows-amd64\\scrim2.json", 'r') as f:
#     my_dict1 = json.load(f)

jsonFiles = [my_dict, my_dict1]
#function to parse data


#use for loop as basic tool to print all player data, similar to siege GG
players = []
excelUserLists = len(excelUsername)
firstMapFormatting = []
#Format output of function to what is necessary to the Basic Stats class
#Confirms that a json file has been added
if len(jsonFiles) > 0:
    [users, kills, deaths, eKills, eDeaths, kOST, hs, mk, trade, clutch, plant, defuse, attackerMain, defenderMain, rounds, operatorKills, operatorDeaths, operatorEntryKills, operatorEntryDeaths, operatorKOST, operatorHS, operatorMKills, operatorTrades, operatorClutch, operatorPlants, operatorDefusal, operatorRounds, mapPlayed, team1Score, team2Score, team1Name, team2Name] = singleMap(jsonFiles[0])
    headerPrint(mapPlayed, team1Score, team2Score, team1Name, team2Name)
    for i in range(len(users)):
        matching = 0
        matchingValue = 0
        #Intermediate values to make a potentially undefined value, or additional info necessary, like entry being subtracted
        #IE, if a person played 1 round, got 2 kills and didnt die, KD would be infinite if not reset
        operatorEntry = np.zeros(opNumbers)
        operatorKD = np.zeros(opNumbers)
        operatorKOSTAmount = np.zeros(opNumbers)
        operatorKPR = np.zeros(opNumbers)
        operatorSRV = np.zeros(opNumbers)
        kd = kills[i]/deaths[i]
        entry = eKills[i]-eDeaths[i]
        kOSTAmount = kOST[i]/rounds[i]
        kPR = kills[i]/rounds[i]
        sRV = (rounds[i]-deaths[i])/rounds[i]
        #Repeats for number of operators as similar stats
        for j in range(opNumbers):
            operatorEntry[j] = operatorEntryKills[i][j] - operatorEntryDeaths[i][j]
            if operatorDeaths[i][j] == 0:
                operatorKD[j] = operatorKills[i][j]
            else:
                operatorKills[i][j]/operatorDeaths[i][j]
            if operatorRounds[i][j] == 0:
                operatorKOSTAmount[j] = 0
                operatorKPR[j] = 0
                operatorSRV[j] = 0
            else:
                operatorKOSTAmount[j] = operatorKOST[i][j]/operatorRounds[i][j]
                operatorKPR[j] = operatorKills[i][j]/operatorRounds[i][j]
                operatorSRV[j] = (operatorRounds[i][j] - operatorDeaths[i][j])/operatorRounds[i][j]
        #Checks these names against the list in Excel
        
        #If doesnt match, add the player to the list
        for j in range(len(excelUsername)):
            #If it matches, can confirm that they have played prior
            if users[i] == opArray[j][0]:
                matchingValue = j
                matching = matching + 1
        if matching == 0:
            array = []
            operatorRating = []
            singleUser = []
            tempKD = 0
            #Same stopping of infinite values and replacing or undefined values
            for b in range(len(operatorsToNum)):
                if operatorKills[i][b] == 0:
                    tempHS = 0
                else:
                    tempHS = operatorHS[i][b]/operatorKills[i][b]
                if operatorDeaths[i][b] == 0:
                    tempKD = operatorKills[i][b]
                else:
                    tempKD = operatorKills[i][b]/operatorDeaths[i][b]
                if operatorRounds[i][b] == 0:
                    tempKOST = 0
                    tempSRV = 0
                    tempRating = 0
                else:
                    tempKOST = operatorKOST[i][b]/operatorRounds[i][b]
                    tempSRV = (operatorRounds[i][b] - operatorDeaths[i][b])/operatorRounds[i][b]
                    tempRating = ratingSys(operatorKills[i][b],tempKD, operatorMKills[i][b], operatorEntryKills[i][b] - operatorEntryDeaths[i][b], operatorPlants[i][b], operatorClutch[i][b], tempKOST, tempSRV, operatorRounds[i][b] )

                #Array to add to excel when all compiled
                array.append(round((tempRating),2))
                operatorRating.append(round((tempRating),2))
                array.append(operatorKills[i][b])
                array.append(operatorDeaths[i][b])
                array.append(operatorEntryKills[i][b])
                array.append(operatorEntryDeaths[i][b])
                array.append(round(tempKOST,2))
                array.append(round(tempHS,2))
                array.append(operatorMKills[i][b])
                array.append(operatorTrades[i][b])
                array.append(operatorClutch[i][b])
                array.append(operatorPlants[i][b])
                array.append(operatorRounds[i][b])
            singleUser.append(users[i])
            #Within the Operators, looking at rounds for the operator plays and split between attackers and defenders
            operatorDefRounds = operatorRounds[i][0:33]
            operatorDefRating = operatorRating[0:33]
            operatorAtkRounds = operatorRounds[i][34:]
            operatorAtkRating = operatorRating[34:]
            
            #Find the max of the Rounds and ratings, these arrays are of slightly different types, but calculate the same thing
            maxDefRounds = np.argmax(operatorDefRounds)
            maxAtkRounds = np.argmax(operatorAtkRounds)
            maxDefRating = max(operatorDefRating)
            maxAtkRating = max(operatorAtkRating)
            
            #Append these values to the user array for excel
            singleUser.append(numToOperators[34+maxAtkRounds])
            singleUser.append(operatorRounds[i][34+maxAtkRounds])
            singleUser.append(numToOperators[maxDefRounds])
            singleUser.append(operatorRounds[i][maxDefRounds])
            
            singleUser.append(numToOperators[34+operatorAtkRating.index(maxAtkRating)])
            singleUser.append(operatorRating[34+operatorAtkRating.index(maxAtkRating)])
            singleUser.append(numToOperators[operatorDefRating.index(maxDefRating)])
            singleUser.append(operatorRating[operatorDefRating.index(maxDefRating)])

            #Adds the operator values to the single User arrays
            for f in range(len(array)):
                singleUser.append(array[f])
            opArray.append(singleUser)
        #This is the values when a player already exists in excel
        else:
            tempAtkRating = []
            tempDefRating = []
            tempAtkRounds = []
            tempDefRounds = []
            #Add values the operator array that will be put into the excel
            #Updates the operator values that already exist in the values
            #Offsets based on excel to access the correct data
            for d in range(len(operatorsToNum)):
                #This is checking headshots and headshot value
                if opArray[matchingValue][10+12*d] == 0:
                    opArray[matchingValue][15+12*d] = 0
                else:
                    opArray[matchingValue][15+12*d] = round((opArray[matchingValue][15+12*d]*opArray[matchingValue][10+12*d] + operatorHS[i][d])/(opArray[matchingValue][10+12*d] + operatorKills[i][d]),2)
                #Updates all values that are empirical
                opArray[matchingValue][10+12*d] = opArray[matchingValue][10+12*d] + operatorKills[i][d]
                opArray[matchingValue][11+12*d] = opArray[matchingValue][11+12*d] + operatorDeaths[i][d]
                opArray[matchingValue][12+12*d] = opArray[matchingValue][12+12*d] + operatorEntryKills[i][d]
                opArray[matchingValue][13+12*d] = opArray[matchingValue][13+12*d] + operatorEntryDeaths[i][d]
                opArray[matchingValue][16+12*d] = opArray[matchingValue][16+12*d] + operatorMKills[i][d]
                opArray[matchingValue][17+12*d] = opArray[matchingValue][17+12*d] + operatorTrades[i][d]
                opArray[matchingValue][18+12*d] = opArray[matchingValue][18+12*d] + operatorClutch[i][d]
                opArray[matchingValue][19+12*d] = opArray[matchingValue][19+12*d] + operatorPlants[i][d]
                opArray[matchingValue][20+12*d] = opArray[matchingValue][20+12*d] + operatorRounds[i][d]
                #Calculates Operator KD
                if opArray[matchingValue][11+12*d] == 0:
                    tempKD = opArray[matchingValue][10+12*d]
                else:
                    tempKD = opArray[matchingValue][10+12*d]/opArray[matchingValue][11+12*d]
                #Calculates survival
                if opArray[matchingValue][20+12*d] == 0:
                    opArray[matchingValue][14+12*d] = 0
                    tempSRV = 0
                    tempRating = 0
                else:
                    opArray[matchingValue][14+12*d] = round((opArray[matchingValue][14+12*d]*opArray[matchingValue][20+12*d] + operatorKOST[i][d])/(opArray[matchingValue][20+12*d] + operatorRounds[i][d]),2)
                    tempSRV = (opArray[matchingValue][20+12*d] - opArray[matchingValue][11+12*d])/opArray[matchingValue][20+12*d]
                    tempRating = round(ratingSys(opArray[matchingValue][10+12*d], tempKD, opArray[matchingValue][16+12*d], opArray[matchingValue][12+12*d] - opArray[matchingValue][13+12*d], opArray[matchingValue][19+12*d], opArray[matchingValue][18+12*d], opArray[matchingValue][14+12*d], tempSRV, opArray[matchingValue][20+12*d]),2)
                opArray[matchingValue][9+12*d] = tempRating
                #Updates the attacker ratings per operator
                if d >= 34:
                    tempAtkRating.append(opArray[matchingValue][9+12*d])
                    
                    
                    tempAtkRounds.append(opArray[matchingValue][20+12*d])
                else:
                    
                    tempDefRating.append(opArray[matchingValue][9+12*d])
                    tempDefRounds.append(opArray[matchingValue][20+12*d])
            #Puts the max ratings and rounds into the array
            maxAtkRating = max(tempAtkRating)
            maxDefRating = max(tempDefRating)
            maxAtkRounds = max(tempAtkRounds)
            maxDefRounds = max(tempDefRounds)
            opArray[matchingValue][1] = numToOperators[34+tempAtkRounds.index(maxAtkRounds)]
            opArray[matchingValue][2] = maxAtkRounds
            opArray[matchingValue][3] = numToOperators[tempDefRounds.index(maxDefRounds)]
            opArray[matchingValue][4] = maxDefRounds
            opArray[matchingValue][5] = numToOperators[34+tempAtkRating.index(maxAtkRating)]
            opArray[matchingValue][6] = maxAtkRating
            opArray[matchingValue][7] = numToOperators[tempDefRating.index(maxDefRating)]
            opArray[matchingValue][8] = maxDefRating

        matching = 0
        matchingValue = -1
        for j in range(len(excelUsername)):
            if excelUsername[j] == users[i]:
                matching = matching + 1
                matchingValue = j
        #If do not match append to the arrays
        if matching == 0:
            excelUsername.append(users[i])
            excelRating.append(round(ratingSys(kills[i], kills[i]/deaths[i], mk[i], eKills[i] - eDeaths[i], plant[i], clutch[i], kOST[i]/rounds[i], (rounds[i]-deaths[i])/rounds[i], rounds[i]),2))
            excelKills.append(kills[i])
            excelDeaths.append(deaths[i])
            excelKD.append(round(kills[i]/deaths[i], 2))
            excelEntryKill.append(eKills[i])
            excelEntryDeath.append(eDeaths[i])
            excelEntryPlusMinus.append(eKills[i] - eDeaths[i])
            excelKOST.append(round(kOST[i]/rounds[i],2))
            excelKPR.append(round(kills[i]/rounds[i],2))
            excelSRV.append(round(((rounds[i]-deaths[i])/rounds[i]),2))
            excelMKills.append(round(mk[i]/rounds[i],2))
            excelTrade.append(round(trade[i]/deaths[i],2))
            excelClutch.append(round(clutch[i]/rounds[i],2))
            excelPlants.append(round(plant[i]/(rounds[i]/2),2))
            excelHS.append(math.ceil(hs[i]))
            excelRound.append(rounds[i])
        #If match, update the value and replace at position
        else:
            excelHS[matchingValue] = math.ceil((excelHS[matchingValue]*excelKills[matchingValue] + hs[i]*kills[i])/(excelKills[matchingValue] + kills[i]))
            excelKills[matchingValue] = excelKills[matchingValue] + kills[i]
            excelDeaths[matchingValue] = excelDeaths[matchingValue] + deaths[i]
            excelKD[matchingValue] = round(excelKills[matchingValue]/excelDeaths[matchingValue],2)
            excelEntryKill[matchingValue] = excelEntryKill[matchingValue] + eKills[i]
            excelEntryDeath[matchingValue] = excelEntryDeath[matchingValue] + eDeaths[i]
            excelEntryPlusMinus[matchingValue] = excelEntryKill[matchingValue] - excelEntryDeath[matchingValue]
            excelKOST[matchingValue] = round((excelKOST[matchingValue]*excelRound[matchingValue] + kOST[i])/(excelRound[matchingValue] + rounds[i]),2)
            excelKPR[matchingValue] = round((excelKPR[matchingValue]*excelRound[matchingValue] + kills[i])/(excelRound[matchingValue] + rounds[i]),2)
            excelSRV[matchingValue] = round(((excelRound[matchingValue] + rounds[i]) - excelDeaths[matchingValue])/(excelRound[matchingValue] + rounds[i]),2)
            excelMKills[matchingValue] = round((excelMKills[matchingValue]*excelRound[matchingValue] + mk[i])/(excelRound[matchingValue]+rounds[i]),2)
            excelTrade[matchingValue] = round((excelTrade[matchingValue]*excelRound[matchingValue] + trade[i])/(excelRound[matchingValue]+rounds[i]),2)
            excelClutch[matchingValue] = round((excelClutch[matchingValue]*excelRound[matchingValue] + clutch[i])/(excelRound[matchingValue]+rounds[i]),2)
            excelPlants[matchingValue] = round((excelPlants[matchingValue]*excelRound[matchingValue] + plant[i])/(excelRound[matchingValue]+rounds[i]),2)
            
            excelRound[matchingValue] = excelRound[matchingValue] + rounds[i]
            excelRating[matchingValue] = round(ratingSys(excelKills[matchingValue], excelKD[matchingValue], excelMKills[matchingValue], excelEntryPlusMinus[matchingValue], excelPlants[matchingValue], excelClutch[matchingValue], excelKOST[matchingValue], excelSRV[matchingValue], excelRound[matchingValue]),2)
    # If only one file is to be read, just add normally
    if len(jsonFiles) == 1:
        for i in range(len(users)):
            kd = kills[i]/deaths[i]
            entry = eKills[i]-eDeaths[i]
            kOSTAmount = kOST[i]/rounds[i]
            kPR = kills[i]/rounds[i]
            sRV = (rounds[i]-deaths[i])/rounds[i]
            player = basicStats(users[i],kills[i],deaths[i],kd,eKills[i],eDeaths[i],entry,kOSTAmount,kPR,sRV,mk[i],trade[i],clutch[i],plant[i],defuse[i],hs[i],attackerMain[i], defenderMain[i], rounds[i],operatorKills[i], operatorDeaths[i],operatorKD, operatorEntryKills[i], operatorEntryDeaths[i], operatorEntry, operatorKOSTAmount, operatorKPR, operatorSRV, operatorMKills[i], operatorTrades[i], operatorClutch[i], operatorPlants[i], operatorDefusal[i], operatorHS[i], operatorRounds[i], mapPlayed, team1Score, team2Score, team1Name, team2Name)
            players.append(player)
            #Prints to terminal
            
            if i == 0:
                start = 1
            else:
                start = 0
            players[i].printIndivStat(start)
    #If more than one json file, then the stats must be combined to read all maps
    else:
        firstMapFormatting = users
        combinedKills = kills
        combinedDeaths = deaths
        combinedEntryKills = eKills
        combinedEntryDeaths = eDeaths
        combinedKOST = kOST
        combinedMultiKills = mk
        combinedClutch = clutch
        combinedTrades = trade
        combinedPlants = plant
        combinedHS = hs
        combinedRounds = rounds
        combinedOpKills = operatorKills
        combinedOpDeaths = operatorDeaths
        combinedOpEntryKills = operatorEntryKills
        combinedOpEntryDeaths = operatorEntryDeaths
        combinedOpKOST = operatorKOST
        combinedOpMultiKills = operatorMKills
        combinedOpTrades = operatorTrades
        combinedOpClutch = operatorClutch
        combinedOpPlant = operatorPlants
        combinedOpHS = operatorHS
        combinedOpRounds = operatorRounds
        wrongPlayersFlag = 0
        #Loop for check to correct error
        # Do not read the first file because it was read originally so read file 2 through whatever
        for j in range(len(jsonFiles[1:])):
            #Reset the map stats for the new map
            [users, kills, deaths, eKills, eDeaths, kOST, hs, mk, trade, clutch, plant, defuse, attackerMain, defenderMain, rounds, operatorKills, operatorDeaths, operatorEntryKills, operatorEntryDeaths, operatorKOST, operatorHS, operatorMKills, operatorTrades, operatorClutch, operatorPlants, operatorDefusal, operatorRounds, mapPlayed, team1Score, team2Score, team1Name, team2Name] = singleMap(jsonFiles[j+1])
            #print the header 
            headerPrint(mapPlayed, team1Score, team2Score, team1Name, team2Name)
            #match the order of the names from different maps, this allows the data to line up to the right person
            #Loop through all users from firstmap formating to the second map
            for k in range(len(users)):
                l = 0
                while l < len(users):
                    #If the name from the first map matches a name in the second, uses position of first for rest of code, but knows where it is on second map
                    if firstMapFormatting[k] == users[l]:
                        break
                    l = l + 1
                if l == 10:
                    wrongPlayersFlag = 1
                    break
                #Combines values to keep updated thru multiple maps
                combinedHS[k] = (hs[l]*kills[l] + combinedHS[k]*combinedKills[k])/(kills[l] + combinedKills[k])
                combinedKills[k] = kills[l] + combinedKills[k]
                combinedDeaths[k] = deaths[l] + combinedDeaths[k]
                combinedEntryKills[k] = eKills[l] + combinedEntryKills[k]
                combinedEntryDeaths[k] = eDeaths[l] + combinedEntryDeaths[k]
                combinedKOST[k] = kOST[l] + combinedKOST[k]
                combinedMultiKills[k] = mk[l] + combinedMultiKills[k]
                combinedClutch[k] = clutch[l] + combinedClutch[k]
                combinedTrades[k] = trade[l] + combinedTrades[k]
                combinedPlants[k] = plant[l] + combinedPlants[k]
                combinedRounds[k] = rounds[l] + combinedRounds[k]
                combinedOpKills[k] = operatorKills[l] + combinedOpKills[k]
                combinedOpDeaths[k] = operatorDeaths[l] + combinedOpDeaths[k]
                combinedOpEntryKills[k] = operatorEntryKills[l] + combinedOpEntryKills[k]
                combinedOpEntryDeaths[k] = operatorEntryDeaths[l] + combinedOpEntryDeaths[k]
                combinedOpKOST[k] = operatorKOST[l] + combinedOpKOST[k]
                combinedOpMultiKills[k] = operatorMKills[l] + combinedOpMultiKills[k]
                combinedOpTrades[k] = operatorTrades[l] + combinedOpTrades[k]
                combinedOpClutch[k] = operatorClutch[l] + combinedOpClutch[k]
                combinedOpPlant[k] = operatorPlants[l] + combinedOpPlant[k]
                combinedOpHS[k] = operatorHS[l] + combinedOpHS[k]
                combinedOpRounds[k] = operatorRounds[l] + combinedOpRounds[k]
            
            
            
            
            #This is the same as above (most likely can be reduced, but havent looked through yet)
            for i in range(len(users)):
                matching = 0
                matchingValue = 0
                #Intermediate values to make a potentially undefined value, or additional info necessary, like entry being subtracted
                #IE, if a person played 1 round, got 2 kills and didnt die, KD would be infinite if not reset
                operatorEntry = np.zeros(opNumbers)
                operatorKD = np.zeros(opNumbers)
                operatorKOSTAmount = np.zeros(opNumbers)
                operatorKPR = np.zeros(opNumbers)
                operatorSRV = np.zeros(opNumbers)
                kd = kills[i]/deaths[i]
                entry = eKills[i]-eDeaths[i]
                kOSTAmount = kOST[i]/rounds[i]
                kPR = kills[i]/rounds[i]
                sRV = (rounds[i]-deaths[i])/rounds[i]
                #Repeats for number of operators as similar stats
                for j in range(opNumbers):
                    operatorEntry[j] = operatorEntryKills[i][j] - operatorEntryDeaths[i][j]
                    if operatorDeaths[i][j] == 0:
                        operatorKD[j] = operatorKills[i][j]
                    else:
                        operatorKills[i][j]/operatorDeaths[i][j]
                    if operatorRounds[i][j] == 0:
                        operatorKOSTAmount[j] = 0
                        operatorKPR[j] = 0
                        operatorSRV[j] = 0
                    else:
                        operatorKOSTAmount[j] = operatorKOST[i][j]/operatorRounds[i][j]
                        operatorKPR[j] = operatorKills[i][j]/operatorRounds[i][j]
                        operatorSRV[j] = (operatorRounds[i][j] - operatorDeaths[i][j])/operatorRounds[i][j]
                #Checks these names against the list in Excel
                
                #If doesnt match, add the player to the list
                for j in range(len(excelUsername)):
                    #If it matches, can confirm that they have played prior
                    if users[i] == opArray[j][0]:
                        matchingValue = j
                        matching = matching + 1
                if matching == 0:
                    array = []
                    operatorRating = []
                    singleUser = []
                    tempKD = 0
                    #Same stopping of infinite values and replacing or undefined values
                    for b in range(len(operatorsToNum)):
                        if operatorKills[i][b] == 0:
                            tempHS = 0
                        else:
                            tempHS = operatorHS[i][b]/operatorKills[i][b]
                        if operatorDeaths[i][b] == 0:
                            tempKD = operatorKills[i][b]
                        else:
                            tempKD = operatorKills[i][b]/operatorDeaths[i][b]
                        if operatorRounds[i][b] == 0:
                            tempKOST = 0
                            tempSRV = 0
                            tempRating = 0
                        else:
                            tempKOST = operatorKOST[i][b]/operatorRounds[i][b]
                            tempSRV = (operatorRounds[i][b] - operatorDeaths[i][b])/operatorRounds[i][b]
                            tempRating = ratingSys(operatorKills[i][b],tempKD, operatorMKills[i][b], operatorEntryKills[i][b] - operatorEntryDeaths[i][b], operatorPlants[i][b], operatorClutch[i][b], tempKOST, tempSRV, operatorRounds[i][b] )

                        #Array to add to excel when all compiled
                        array.append(round((tempRating),2))
                        operatorRating.append(round((tempRating),2))
                        array.append(operatorKills[i][b])
                        array.append(operatorDeaths[i][b])
                        array.append(operatorEntryKills[i][b])
                        array.append(operatorEntryDeaths[i][b])
                        array.append(round(tempKOST,2))
                        array.append(round(tempHS,2))
                        array.append(operatorMKills[i][b])
                        array.append(operatorTrades[i][b])
                        array.append(operatorClutch[i][b])
                        array.append(operatorPlants[i][b])
                        array.append(operatorRounds[i][b])
                    singleUser.append(users[i])
                    #Within the Operators, looking at rounds for the operator plays and split between attackers and defenders
                    operatorDefRounds = operatorRounds[i][0:33]
                    operatorDefRating = operatorRating[0:33]
                    operatorAtkRounds = operatorRounds[i][34:]
                    operatorAtkRating = operatorRating[34:]
                    
                    #Find the max of the Rounds and ratings, these arrays are of slightly different types, but calculate the same thing
                    maxDefRounds = np.argmax(operatorDefRounds)
                    maxAtkRounds = np.argmax(operatorAtkRounds)
                    maxDefRating = max(operatorDefRating)
                    maxAtkRating = max(operatorAtkRating)
                    
                    #Append these values to the user array for excel
                    singleUser.append(numToOperators[34+maxAtkRounds])
                    singleUser.append(operatorRounds[i][34+maxAtkRounds])
                    singleUser.append(numToOperators[maxDefRounds])
                    singleUser.append(operatorRounds[i][maxDefRounds])
                    
                    singleUser.append(numToOperators[34+operatorAtkRating.index(maxAtkRating)])
                    singleUser.append(operatorRating[34+operatorAtkRating.index(maxAtkRating)])
                    singleUser.append(numToOperators[operatorDefRating.index(maxDefRating)])
                    singleUser.append(operatorRating[operatorDefRating.index(maxDefRating)])

                    #Adds the operator values to the single User arrays
                    for f in range(len(array)):
                        singleUser.append(array[f])
                    opArray.append(singleUser)
                #This is the values when a player already exists in excel
                else:
                    tempAtkRating = []
                    tempDefRating = []
                    tempAtkRounds = []
                    tempDefRounds = []
                    #Add values the operator array that will be put into the excel
                    #Updates the operator values that already exist in the values
                    #Offsets based on excel to access the correct data
                    for d in range(len(operatorsToNum)):
                        #This is checking headshots and headshot value
                        if opArray[matchingValue][10+12*d] == 0:
                            opArray[matchingValue][15+12*d] = 0
                        else:
                            opArray[matchingValue][15+12*d] = round((opArray[matchingValue][15+12*d]*opArray[matchingValue][10+12*d] + operatorHS[i][d])/(opArray[matchingValue][10+12*d] + operatorKills[i][d]),2)
                        #Updates all values that are empirical
                        opArray[matchingValue][10+12*d] = opArray[matchingValue][10+12*d] + operatorKills[i][d]
                        opArray[matchingValue][11+12*d] = opArray[matchingValue][11+12*d] + operatorDeaths[i][d]
                        opArray[matchingValue][12+12*d] = opArray[matchingValue][12+12*d] + operatorEntryKills[i][d]
                        opArray[matchingValue][13+12*d] = opArray[matchingValue][13+12*d] + operatorEntryDeaths[i][d]
                        opArray[matchingValue][16+12*d] = opArray[matchingValue][16+12*d] + operatorMKills[i][d]
                        opArray[matchingValue][17+12*d] = opArray[matchingValue][17+12*d] + operatorTrades[i][d]
                        opArray[matchingValue][18+12*d] = opArray[matchingValue][18+12*d] + operatorClutch[i][d]
                        opArray[matchingValue][19+12*d] = opArray[matchingValue][19+12*d] + operatorPlants[i][d]
                        opArray[matchingValue][20+12*d] = opArray[matchingValue][20+12*d] + operatorRounds[i][d]
                        #Calculates Operator KD
                        if opArray[matchingValue][11+12*d] == 0:
                            tempKD = opArray[matchingValue][10+12*d]
                        else:
                            tempKD = opArray[matchingValue][10+12*d]/opArray[matchingValue][11+12*d]
                        #Calculates survival
                        if opArray[matchingValue][20+12*d] == 0:
                            opArray[matchingValue][14+12*d] = 0
                            tempSRV = 0
                            tempRating = 0
                        else:
                            opArray[matchingValue][14+12*d] = round((opArray[matchingValue][14+12*d]*opArray[matchingValue][20+12*d] + operatorKOST[i][d])/(opArray[matchingValue][20+12*d] + operatorRounds[i][d]),2)
                            tempSRV = (opArray[matchingValue][20+12*d] - opArray[matchingValue][11+12*d])/opArray[matchingValue][20+12*d]
                            tempRating = round(ratingSys(opArray[matchingValue][10+12*d], tempKD, opArray[matchingValue][16+12*d], opArray[matchingValue][12+12*d] - opArray[matchingValue][13+12*d], opArray[matchingValue][19+12*d], opArray[matchingValue][18+12*d], opArray[matchingValue][14+12*d], tempSRV, opArray[matchingValue][20+12*d]),2)
                        opArray[matchingValue][9+12*d] = tempRating
                        #Updates the attacker ratings per operator
                        if d >= 34:
                            tempAtkRating.append(opArray[matchingValue][9+12*d])
                            
                            
                            tempAtkRounds.append(opArray[matchingValue][20+12*d])
                        else:
                            
                            tempDefRating.append(opArray[matchingValue][9+12*d])
                            tempDefRounds.append(opArray[matchingValue][20+12*d])
                    #Puts the max ratings and rounds into the array
                    maxAtkRating = max(tempAtkRating)
                    maxDefRating = max(tempDefRating)
                    maxAtkRounds = max(tempAtkRounds)
                    maxDefRounds = max(tempDefRounds)
                    opArray[matchingValue][1] = numToOperators[34+tempAtkRounds.index(maxAtkRounds)]
                    opArray[matchingValue][2] = maxAtkRounds
                    opArray[matchingValue][3] = numToOperators[tempDefRounds.index(maxDefRounds)]
                    opArray[matchingValue][4] = maxDefRounds
                    opArray[matchingValue][5] = numToOperators[34+tempAtkRating.index(maxAtkRating)]
                    opArray[matchingValue][6] = maxAtkRating
                    opArray[matchingValue][7] = numToOperators[tempDefRating.index(maxDefRating)]
                    opArray[matchingValue][8] = maxDefRating

                matching = 0
                matchingValue = -1
                for j in range(len(excelUsername)):
                    if excelUsername[j] == users[i]:
                        matching = matching + 1
                        matchingValue = j
                #If do not match append to the arrays
                if matching == 0:
                    excelUsername.append(users[i])
                    excelRating.append(round(ratingSys(kills[i], kills[i]/deaths[i], mk[i], eKills[i] - eDeaths[i], plant[i], clutch[i], kOST[i]/rounds[i], (rounds[i]-deaths[i])/rounds[i], rounds[i]),2))
                    excelKills.append(kills[i])
                    excelDeaths.append(deaths[i])
                    excelKD.append(round(kills[i]/deaths[i], 2))
                    excelEntryKill.append(eKills[i])
                    excelEntryDeath.append(eDeaths[i])
                    excelEntryPlusMinus.append(eKills[i] - eDeaths[i])
                    excelKOST.append(round(kOST[i]/rounds[i],2))
                    excelKPR.append(round(kills[i]/rounds[i],2))
                    excelSRV.append(round(((rounds[i]-deaths[i])/rounds[i]),2))
                    excelMKills.append(round(mk[i]/rounds[i],2))
                    excelTrade.append(round(trade[i]/deaths[i],2))
                    excelClutch.append(round(clutch[i]/rounds[i],2))
                    excelPlants.append(round(plant[i]/(rounds[i]/2),2))
                    excelHS.append(math.ceil(hs[i]))
                    excelRound.append(rounds[i])
                #If match, update the value and replace at position
                else:
                    excelHS[matchingValue] = math.ceil((excelHS[matchingValue]*excelKills[matchingValue] + hs[i]*kills[i])/(excelKills[matchingValue] + kills[i]))
                    excelKills[matchingValue] = excelKills[matchingValue] + kills[i]
                    excelDeaths[matchingValue] = excelDeaths[matchingValue] + deaths[i]
                    excelKD[matchingValue] = round(excelKills[matchingValue]/excelDeaths[matchingValue],2)
                    excelEntryKill[matchingValue] = excelEntryKill[matchingValue] + eKills[i]
                    excelEntryDeath[matchingValue] = excelEntryDeath[matchingValue] + eDeaths[i]
                    excelEntryPlusMinus[matchingValue] = excelEntryKill[matchingValue] - excelEntryDeath[matchingValue]
                    excelKOST[matchingValue] = round((excelKOST[matchingValue]*excelRound[matchingValue] + kOST[i])/(excelRound[matchingValue] + rounds[i]),2)
                    excelKPR[matchingValue] = round((excelKPR[matchingValue]*excelRound[matchingValue] + kills[i])/(excelRound[matchingValue] + rounds[i]),2)
                    excelSRV[matchingValue] = round(((excelRound[matchingValue] + rounds[i]) - excelDeaths[matchingValue])/(excelRound[matchingValue] + rounds[i]),2)
                    excelMKills[matchingValue] = round((excelMKills[matchingValue]*excelRound[matchingValue] + mk[i])/(excelRound[matchingValue]+rounds[i]),2)
                    excelTrade[matchingValue] = round((excelTrade[matchingValue]*excelRound[matchingValue] + trade[i])/(excelRound[matchingValue]+rounds[i]),2)
                    excelClutch[matchingValue] = round((excelClutch[matchingValue]*excelRound[matchingValue] + clutch[i])/(excelRound[matchingValue]+rounds[i]),2)
                    excelPlants[matchingValue] = round((excelPlants[matchingValue]*excelRound[matchingValue] + plant[i])/(excelRound[matchingValue]+rounds[i]),2)
                    
                    excelRound[matchingValue] = excelRound[matchingValue] + rounds[i]
                    excelRating[matchingValue] = round(ratingSys(excelKills[matchingValue], excelKD[matchingValue], excelMKills[matchingValue], excelEntryPlusMinus[matchingValue], excelPlants[matchingValue], excelClutch[matchingValue], excelKOST[matchingValue], excelSRV[matchingValue], excelRound[matchingValue]),2)
        #check for all users combined to set values up to be sent to the match stats printout
        for i in range(len(users)):
            if combinedDeaths[i] == 0:
                combinedKD = combinedKills[i]
            else:
                combinedKD = combinedKills[i]/combinedDeaths[i]
            #!!! Need to do def and atk mains
            combinedEntry = combinedEntryKills[i] - combinedEntryDeaths[i]
            combinedKOSTAmount = combinedKOST[i]/combinedRounds[i]
            combinedKPR = combinedKills[i]/combinedRounds[i]
            combinedSRV = (combinedRounds[i] - combinedDeaths[i])/combinedRounds[i] 
            
            player = basicStats(firstMapFormatting[i],combinedKills[i],combinedDeaths[i],combinedKD,combinedEntryKills[i],combinedEntryDeaths[i],combinedEntry,combinedKOSTAmount,combinedKPR,combinedSRV,combinedMultiKills[i],combinedTrades[i],combinedClutch[i],combinedPlants[i],defuse[i],combinedHS[i],attackerMain[i], defenderMain[i], rounds[i],operatorKills[i], operatorDeaths[i],operatorKD, operatorEntryKills[i], operatorEntryDeaths[i], operatorEntry, operatorKOSTAmount, operatorKPR, operatorSRV, operatorMKills[i], operatorTrades[i], operatorClutch[i], operatorPlants[i], operatorDefusal[i], operatorHS[i], operatorRounds[i], mapPlayed, team1Score, team2Score, team1Name, team2Name)
            players.append(player)
            if i == 0:
                start = 1
            else:
                start = 0
            players[i].printIndivStat(start)
else:
    print("ERROR: No json files have been selected")

#Can be an outdated value

#For the number of users in the current match
    

    
            
            



#Check new or returning players for the Stats Excel Page
for column_index, value in enumerate(operatorHeader, start=1):
    cell = excelOpStatSheet.cell(row=1, column=column_index)
    cell.value = value

for row_index, row in enumerate(opArray, start=2):
    for column_index, value in enumerate(row, start=1):
        cell = excelOpStatSheet.cell(row=row_index, column=column_index)
        cell.value = value


for i in range(len(excelArr)):
    for j in range(len(excelArr[i])):
        cell = excelMainSheet[excelCols[i] + str(j+2)]
        cell.value = excelArr[i][j]

if wrongPlayersFlag != 0:
    print('ERROR: Not same 10 players seen through multiple maps')
# Save the changes
#This is how to get the Excel file to save the changes that were made
#workbook.save(filename=excelFile)


#These are examples of how to access the data in python per operator stats
#players[9].SingleOperatorStats('Flores')
#players[7].allOps()








